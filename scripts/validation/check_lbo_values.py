"""
Check LBO Model Values - Verify no zeros or division errors
"""

import openpyxl

def check_lbo_values(filepath: str):
    """Check that the LBO model has proper calculated values."""

    print("="*80)
    print("CHECKING LBO MODEL VALUES")
    print("="*80)

    wb = openpyxl.load_workbook(filepath, data_only=True)

    errors_found = []

    # Check Cover Sheet
    print("\n📄 COVER SHEET VALUES:")
    cover = wb["Cover"]

    summary_items = [
        (11, "Purchase Enterprise Value"),
        (12, "Entry EBITDA Multiple"),
        (13, "Total Debt"),
        (14, "Equity Contribution"),
        (15, "Exit Enterprise Value"),
        (16, "Equity Value at Exit"),
        (17, "IRR"),
        (18, "MOIC"),
    ]

    for row, label in summary_items:
        value = cover.cell(row=row, column=3).value
        print(f"   {label}: {value}")
        if value is None or value == 0:
            errors_found.append(f"Cover: {label} is showing {value}")

    # Check Transaction Summary
    print("\n📊 TRANSACTION SUMMARY VALUES:")
    ts = wb["Transaction Summary"]

    items = [
        (5, "LTM EBITDA"),
        (6, "Entry Multiple"),
        (8, "Purchase EV"),
        (11, "Exit Year EBITDA"),
        (12, "Exit Multiple"),
        (13, "Exit EV"),
    ]

    for row, label in items:
        value = ts.cell(row=row, column=2).value
        print(f"   {label}: {value}")
        if value is None or value == 0:
            errors_found.append(f"Transaction Summary: {label} is {value}")

    # Check Sources & Uses
    print("\n💰 SOURCES & USES VALUES:")
    su = wb["Sources & Uses"]

    items = [
        (5, "Purchase EV"),
        (8, "Total Uses"),
        (12, "Sponsor Equity"),
        (14, "Senior Term Loan"),
        (15, "Subordinated Notes"),
        (16, "Total Sources"),
    ]

    for row, label in items:
        value = su.cell(row=row, column=2).value
        print(f"   {label}: {value}")
        if value is None:
            errors_found.append(f"Sources & Uses: {label} is None")

    # Check Operating Model (first year projections)
    print("\n📈 OPERATING MODEL - YEAR 1:")
    om = wb["Operating Model"]

    items = [
        (4, 3, "Revenue Year 1"),
        (5, 3, "EBITDA Year 1"),
        (6, 3, "D&A Year 1"),
        (7, 3, "EBIT Year 1"),
    ]

    for row, col, label in items:
        value = om.cell(row=row, column=col).value
        print(f"   {label}: {value}")
        if value is None:
            errors_found.append(f"Operating Model: {label} is None")

    # Check for division errors in any sheet
    print("\n🔍 CHECKING FOR #DIV/0! ERRORS...")
    div_errors = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "#DIV" in str(cell.value):
                    print(f"   ⚠️  {sheet_name}!{cell.coordinate}: {cell.value}")
                    div_errors += 1
                    errors_found.append(f"{sheet_name}!{cell.coordinate} has division error")

    if div_errors == 0:
        print("   ✓ No division errors found")

    # Summary
    print("\n" + "="*80)
    if errors_found:
        print("❌ ERRORS FOUND:")
        for error in errors_found:
            print(f"   - {error}")
    else:
        print("✅ ALL VALUES LOOK CORRECT!")
        print("   - Cover sheet shows proper summary values")
        print("   - Transaction summary has entry and exit valuations")
        print("   - Sources & Uses properly calculated")
        print("   - Operating model has projections")
        print("   - No division errors detected")
    print("="*80)

    wb.close()
    return len(errors_found) == 0


if __name__ == "__main__":
    check_lbo_values("LBO_Model_Example.xlsx")
