"""
Check LBO Model Formulas - Show actual formulas (not values)
"""

import openpyxl

def check_lbo_formulas(filepath: str):
    """Check that the LBO model has correct formulas."""

    print("="*80)
    print("CHECKING LBO MODEL FORMULAS")
    print("="*80)

    wb = openpyxl.load_workbook(filepath, data_only=False)

    # Check Cover Sheet formulas
    print("\n📄 COVER SHEET FORMULAS:")
    cover = wb["Cover"]

    summary_items = [
        (11, "Purchase Enterprise Value"),
        (12, "Entry EBITDA Multiple"),
        (13, "Total Debt"),
        (14, "Equity Contribution"),
        (15, "Exit Enterprise Value"),
        (16, "Equity Value at Exit"),
        (17, "IRR"),
        (18, "MOIC"),
    ]

    for row, label in summary_items:
        value = cover.cell(row=row, column=3).value
        print(f"   Row {row} - {label}:")
        print(f"      Formula: {value}")

    # Check Transaction Summary
    print("\n📊 TRANSACTION SUMMARY FORMULAS:")
    ts = wb["Transaction Summary"]

    print(f"\n   First 15 rows:")
    for row in range(1, 16):
        col1 = ts.cell(row=row, column=1).value
        col2 = ts.cell(row=row, column=2).value
        if col1 or col2:
            print(f"   Row {row}: {col1} = {col2}")

    # Check Assumptions layout
    print("\n⚙️  ASSUMPTIONS LAYOUT:")
    assump = wb["Assumptions"]

    print(f"\n   First 35 rows:")
    for row in range(1, 36):
        col1 = assump.cell(row=row, column=1).value
        col2 = assump.cell(row=row, column=2).value
        if col1 or col2:
            print(f"   Row {row}: {col1} = {col2}")

    # Check Operating Model layout
    print("\n📈 OPERATING MODEL LAYOUT:")
    om = wb["Operating Model"]

    print(f"\n   First 12 rows, first 5 columns:")
    for row in range(1, 13):
        row_data = []
        for col in range(1, 6):
            val = om.cell(row=row, column=col).value
            row_data.append(str(val) if val else "")
        print(f"   Row {row}: {' | '.join(row_data)}")

    wb.close()


if __name__ == "__main__":
    check_lbo_formulas("Examples/LBO_Model_AcmeTech.xlsx")
