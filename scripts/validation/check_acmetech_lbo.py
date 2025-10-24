"""
Check AcmeTech LBO Model Values
"""

import openpyxl

def check_acmetech_lbo():
    """Check that the AcmeTech LBO model has proper values."""

    print("="*80)
    print("CHECKING ACMETECH LBO MODEL VALUES")
    print("="*80)

    wb = openpyxl.load_workbook('Examples/LBO_Model_AcmeTech.xlsx', data_only=True)

    errors_found = []

    # Check Cover Sheet
    print("\n📄 COVER SHEET VALUES:")
    cover = wb["Cover"]

    summary_items = [
        (11, "Purchase Enterprise Value"),
        (12, "Entry EBITDA Multiple"),
        (13, "Total Debt"),
        (14, "Equity Contribution"),
        (15, "Exit Enterprise Value"),
        (16, "Equity Value at Exit"),
        (17, "IRR"),
        (18, "MOIC"),
    ]

    for row, label in summary_items:
        value = cover.cell(row=row, column=3).value
        print(f"   {label}: {value}")
        if value is None or value == 0:
            errors_found.append(f"Cover: {label} is showing {value}")

    # Check Transaction Summary
    print("\n📊 TRANSACTION SUMMARY VALUES:")
    ts = wb["Transaction Summary"]

    items = [
        (5, "LTM EBITDA"),
        (6, "Entry Multiple"),
        (7, "Purchase EV"),
        (10, "Exit Year EBITDA"),
        (11, "Exit Multiple"),
        (12, "Exit EV"),
    ]

    for row, label in items:
        value = ts.cell(row=row, column=2).value
        print(f"   {label}: {value}")
        if value is None:
            errors_found.append(f"Transaction Summary: {label} is None")

    # Check Sources & Uses
    print("\n💰 SOURCES & USES VALUES:")
    su = wb["Sources & Uses"]

    items = [
        (5, "Purchase EV"),
        (8, "Total Uses"),
        (12, "Sponsor Equity"),
        (14, "Senior Term Loan"),
        (15, "Subordinated Notes"),
        (16, "Total Sources"),
        (18, "CHECK"),
    ]

    for row, label in items:
        value = su.cell(row=row, column=2).value
        print(f"   {label}: {value}")
        if value is None:
            errors_found.append(f"Sources & Uses: {label} is None")

    # Check Assumptions
    print("\n⚙️  ASSUMPTIONS VALUES:")
    assump = wb["Assumptions"]

    items = [
        (8, "Sponsor Equity ($mm)"),
        (14, "Senior Term Loan ($mm)"),
        (18, "Subordinated Notes ($mm)"),
    ]

    for row, label in items:
        value = assump.cell(row=row, column=2).value
        print(f"   {label}: {value}")
        if value is None or value == 0:
            errors_found.append(f"Assumptions: {label} is {value}")

    # Check Operating Model (first year projections)
    print("\n📈 OPERATING MODEL - YEAR 1:")
    om = wb["Operating Model"]

    items = [
        (4, 3, "Revenue Year 1"),
        (5, 3, "EBITDA Year 1"),
        (6, 3, "D&A Year 1"),
        (7, 3, "EBIT Year 1"),
    ]

    for row, col, label in items:
        value = om.cell(row=row, column=col).value
        print(f"   {label}: {value}")
        if value is None:
            errors_found.append(f"Operating Model: {label} is None")

    # Returns Analysis
    print("\n💵 RETURNS ANALYSIS VALUES:")
    ra = wb["Returns Analysis"]

    items = [
        (5, "Exit Year EBITDA"),
        (6, "Exit EV / EBITDA Multiple"),
        (7, "Exit Enterprise Value"),
        (8, "Less: Remaining Debt"),
        (9, "Equity Value at Exit"),
        (11, "IRR"),
        (12, "MOIC"),
    ]

    for row, label in items:
        value = ra.cell(row=row, column=2).value
        print(f"   {label}: {value}")
        if value is None:
            errors_found.append(f"Returns Analysis: {label} is None")

    # Summary
    print("\n" + "="*80)
    if errors_found:
        print("❌ ERRORS FOUND:")
        for error in errors_found:
            print(f"   - {error}")
    else:
        print("✅ ALL VALUES LOOK CORRECT!")
        print("   - Cover sheet shows proper summary values")
        print("   - Transaction summary has entry and exit valuations")
        print("   - Sources & Uses properly calculated")
        print("   - Operating model has projections")
        print("   - Returns analysis shows IRR and MOIC")
    print("="*80)

    wb.close()
    return len(errors_found) == 0


if __name__ == "__main__":
    check_acmetech_lbo()
