"""
Diagnose LBO Model Issues
"""

import openpyxl

def diagnose_lbo_model(filepath: str):
    """Check for errors in the LBO model."""

    print("="*80)
    print("DIAGNOSING LBO MODEL ISSUES")
    print("="*80)

    wb = openpyxl.load_workbook(filepath, data_only=False)

    # Check Cover Sheet
    print("\n📄 COVER SHEET:")
    cover = wb["Cover"]
    print(f"\nFirst 20 rows of Cover sheet:")
    for row in range(1, 21):
        col1 = cover.cell(row=row, column=1).value
        col2 = cover.cell(row=row, column=2).value
        if col1 or col2:
            print(f"   Row {row}: [{col1}] | [{col2}]")

    # Check Transaction Summary
    print("\n📊 TRANSACTION SUMMARY:")
    ts = wb["Transaction Summary"]
    print(f"\nFirst 30 rows:")
    for row in range(1, 31):
        col1 = ts.cell(row=row, column=1).value
        col2 = ts.cell(row=row, column=2).value
        if col1 or col2:
            val_display = col2 if col2 else ""
            print(f"   Row {row}: {col1} = {val_display}")

    # Check Sources & Uses
    print("\n💰 SOURCES & USES:")
    su = wb["Sources & Uses"]
    print(f"\nFirst 40 rows:")
    for row in range(1, 41):
        col1 = su.cell(row=row, column=1).value
        col2 = su.cell(row=row, column=2).value
        if col1 or col2:
            val_display = col2 if col2 else ""
            print(f"   Row {row}: {col1} = {val_display}")

    # Check Assumptions
    print("\n⚙️  ASSUMPTIONS:")
    assump = wb["Assumptions"]
    print(f"\nFirst 40 rows:")
    for row in range(1, 41):
        col1 = assump.cell(row=row, column=1).value
        col2 = assump.cell(row=row, column=2).value
        if col1 or col2:
            val_display = col2 if col2 else ""
            print(f"   Row {row}: {col1} = {val_display}")

    # Check Operating Model
    print("\n📈 OPERATING MODEL:")
    om = wb["Operating Model"]
    print(f"\nFirst 30 rows, first 4 columns:")
    for row in range(1, 31):
        col1 = om.cell(row=row, column=1).value
        col2 = om.cell(row=row, column=2).value
        col3 = om.cell(row=row, column=3).value
        col4 = om.cell(row=row, column=4).value
        if col1 or col2 or col3 or col4:
            print(f"   Row {row}: {col1} | {col2} | {col3} | {col4}")

    wb.close()


if __name__ == "__main__":
    diagnose_lbo_model("LBO_Model_Example.xlsx")
