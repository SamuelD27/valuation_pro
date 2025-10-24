"""
Inspect Financial_Model_Data_Source.xlsx to understand its structure
"""

import openpyxl

def inspect_datasource():
    """Inspect the new data source file."""

    print("="*80)
    print("INSPECTING Financial_Model_Data_Source.xlsx")
    print("="*80)

    wb = openpyxl.load_workbook('Base_datasource/Financial_Model_Data_Source.xlsx')

    print(f"\n📊 SHEETS FOUND: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n" + "="*80)
        print(f"SHEET: {sheet_name}")
        print("="*80)

        # Show first 30 rows and 10 columns
        print(f"\nFirst 30 rows:")
        for row in range(1, min(31, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(11, ws.max_column + 1)):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    row_data.append(f"[{col}]{val}")
            if row_data:
                print(f"   Row {row}: {' | '.join(row_data)}")

    wb.close()


if __name__ == "__main__":
    inspect_datasource()
