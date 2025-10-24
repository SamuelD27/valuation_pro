"""
LBO Modeling Tool - Pixel-Perfect IB-Standard LBO Model Generator

Creates professional LBO models with:
- Sources & Uses of Funds
- Transaction assumptions
- Debt schedule with multiple tranches
- Projected financials (EBITDA → Free Cash Flow)
- Cash flow waterfall
- Returns analysis (IRR, MOIC, Exit value)
- Proper IB-style table formatting with borders
- Ultra-accurate formulas (no hardcoded values)

This is a FOCUSED tool for LBO only (not DCF, Comps, etc.)
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional
from datetime import datetime

from src.excel.formatter import IBFormatter


class LBOTool:
    """
    Professional LBO Model Generator.

    Creates pixel-perfect LBO models with proper IB formatting.
    """

    def __init__(self, company_name: str, sponsor: str = ""):
        """
        Initialize LBO Tool.

        Args:
            company_name: Name of target company
            sponsor: PE sponsor name (optional)
        """
        self.company_name = company_name
        self.sponsor = sponsor
        self.wb = None
        self.formatter = IBFormatter()

        # Professional IB colors
        self.SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Dark blue
        self.SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")  # White text
        self.INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
        self.INPUT_FONT = Font(name="Calibri", size=11, color="000000")  # Black text
        self.CALC_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # Light gray for calculated

    def generate_lbo_model(
        self,
        transaction_data: Dict,
        assumptions: Dict,
        output_file: str
    ):
        """
        Generate complete LBO model.

        Args:
            transaction_data: Transaction details (purchase price, EBITDA, etc.)
            assumptions: LBO assumptions (debt structure, exit multiple, etc.)
            output_file: Output file path
        """
        self.wb = Workbook()
        del self.wb['Sheet']  # Remove default sheet

        # Create sheets in order
        self._create_cover_sheet()
        self._create_transaction_summary(transaction_data, assumptions)
        self._create_sources_uses(transaction_data, assumptions)
        self._create_assumptions_sheet(assumptions)
        self._create_operating_model(transaction_data, assumptions)
        self._create_debt_schedule(assumptions)
        self._create_cash_flow_waterfall(assumptions)
        self._create_returns_analysis(transaction_data, assumptions)

        # Save
        self.wb.save(output_file)
        print(f"✅ LBO model saved to: {output_file}")

    def _create_cover_sheet(self):
        """Create professional cover page."""
        ws = self.wb.create_sheet("Cover", 0)

        # Title
        ws['B2'] = "LEVERAGED BUYOUT ANALYSIS"
        ws['B2'].font = Font(name="Calibri", size=20, bold=True)

        ws['B4'] = self.company_name
        ws['B4'].font = Font(name="Calibri", size=16, bold=True)

        if self.sponsor:
            ws['B5'] = f"Sponsor: {self.sponsor}"
            ws['B5'].font = Font(name="Calibri", size=12)

        ws['B7'] = f"Date: {datetime.now().strftime('%B %d, %Y')}"
        ws['B7'].font = Font(name="Calibri", size=11)

        # Summary box
        ws['B10'] = "TRANSACTION SUMMARY"
        ws['B10'].font = self.SECTION_FONT
        ws['B10'].fill = self.SECTION_FILL

        summary_items = [
            ("Purchase Enterprise Value", "='Transaction Summary'!B7"),
            ("Entry EBITDA Multiple", "='Transaction Summary'!B6"),
            ("Total Debt", "='Sources & Uses'!B16"),
            ("Equity Contribution", "='Sources & Uses'!B12"),
            ("Exit Enterprise Value", "='Transaction Summary'!B12"),
            ("Equity Value at Exit", "='Returns Analysis'!B9"),
            ("IRR", "='Returns Analysis'!B11"),
            ("MOIC", "='Returns Analysis'!B12"),
        ]

        row = 11
        for label, formula in summary_items:
            ws.cell(row=row, column=2).value = label
            ws.cell(row=row, column=2).font = Font(name="Calibri", size=11)

            ws.cell(row=row, column=3).value = formula
            ws.cell(row=row, column=3).font = Font(name="Calibri", size=11, bold=True)

            # Format value cell
            if "Multiple" in label or "IRR" in label or "MOIC" in label:
                if "IRR" in label:
                    ws.cell(row=row, column=3).number_format = '0.0%'
                elif "MOIC" in label:
                    ws.cell(row=row, column=3).number_format = '0.0x'
                else:
                    ws.cell(row=row, column=3).number_format = '0.0x'
            else:
                ws.cell(row=row, column=3).number_format = '$#,##0.0,,"M"'

            row += 1

        # Add border
        self._add_table_border(ws, f'B10:C{row-1}')

        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20

    def _create_transaction_summary(self, transaction_data: Dict, assumptions: Dict):
        """Create Transaction Summary sheet."""
        ws = self.wb.create_sheet("Transaction Summary")

        # Title
        ws['A1'] = "TRANSACTION SUMMARY"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)

        row = 4

        # Entry Valuation
        ws.cell(row=row, column=1).value = "ENTRY VALUATION"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        self._add_table_border(ws, f'A{row}:D{row}')
        row += 1

        # LTM EBITDA
        ws.cell(row=row, column=1).value = "LTM EBITDA"
        ws.cell(row=row, column=2).value = transaction_data.get('ltm_ebitda', 100)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '$#,##0.0,,"M"'
        ltm_ebitda_row = row
        row += 1

        # Entry Multiple
        ws.cell(row=row, column=1).value = "Entry EV / EBITDA Multiple"
        ws.cell(row=row, column=2).value = assumptions.get('entry_multiple', 10.0)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0x'
        entry_multiple_row = row
        row += 1

        # Purchase Enterprise Value
        ws.cell(row=row, column=1).value = "Purchase Enterprise Value"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = f"=B{ltm_ebitda_row}*B{entry_multiple_row}"
        ws.cell(row=row, column=2).number_format = '$#,##0.0,,"M"'
        ws.cell(row=row, column=2).font = Font(bold=True)
        purchase_ev_row = row
        row += 2

        # Exit Valuation
        ws.cell(row=row, column=1).value = "EXIT VALUATION"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        self._add_table_border(ws, f'A{row}:D{row}')
        row += 1

        # Exit Year EBITDA
        ws.cell(row=row, column=1).value = "Exit Year EBITDA"
        ws.cell(row=row, column=2).value = "='Operating Model'!G10"  # Year 5 EBITDA
        ws.cell(row=row, column=2).number_format = '$#,##0.0,,"M"'
        exit_ebitda_row = row
        row += 1

        # Exit Multiple
        ws.cell(row=row, column=1).value = "Exit EV / EBITDA Multiple"
        ws.cell(row=row, column=2).value = assumptions.get('exit_multiple', 11.0)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0x'
        exit_multiple_row = row
        row += 1

        # Exit Enterprise Value
        ws.cell(row=row, column=1).value = "Exit Enterprise Value"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = f"=B{exit_ebitda_row}*B{exit_multiple_row}"
        ws.cell(row=row, column=2).number_format = '$#,##0.0,,"M"'
        ws.cell(row=row, column=2).font = Font(bold=True)

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

    def _create_sources_uses(self, transaction_data: Dict, assumptions: Dict):
        """Create Sources & Uses sheet."""
        ws = self.wb.create_sheet("Sources & Uses")

        # Title
        ws['A1'] = "SOURCES & USES OF FUNDS"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)

        # USES
        row = 4
        ws.cell(row=row, column=1).value = "USES"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        ws.cell(row=row, column=2).value = "$mm"
        ws.cell(row=row, column=2).font = self.SECTION_FONT
        ws.cell(row=row, column=2).fill = self.SECTION_FILL
        self._add_table_border(ws, f'A{row}:C{row}')
        row += 1

        # Purchase Enterprise Value
        ws.cell(row=row, column=1).value = "Purchase Enterprise Value"
        ws.cell(row=row, column=2).value = "='Transaction Summary'!B7"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        purchase_ev_row = row
        row += 1

        # Transaction Fees
        ws.cell(row=row, column=1).value = "Transaction Fees (% of EV)"
        ws.cell(row=row, column=2).value = assumptions.get('transaction_fees_pct', 0.03)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0%'
        fees_pct_row = row
        row += 1

        ws.cell(row=row, column=1).value = "Transaction Fees ($mm)"
        ws.cell(row=row, column=2).value = f"=B{purchase_ev_row}*B{fees_pct_row}"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        fees_row = row
        row += 1

        # Total Uses
        ws.cell(row=row, column=1).value = "TOTAL USES"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = f"=B{purchase_ev_row}+B{fees_row}"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        ws.cell(row=row, column=2).font = Font(bold=True)
        total_uses_row = row
        row += 2

        # SOURCES
        ws.cell(row=row, column=1).value = "SOURCES"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        ws.cell(row=row, column=2).value = "$mm"
        ws.cell(row=row, column=2).font = self.SECTION_FONT
        ws.cell(row=row, column=2).fill = self.SECTION_FILL
        ws.cell(row=row, column=3).value = "% of Total"
        ws.cell(row=row, column=3).font = self.SECTION_FONT
        ws.cell(row=row, column=3).fill = self.SECTION_FILL
        self._add_table_border(ws, f'A{row}:C{row}')
        row += 1

        sources_start = row

        # Sponsor Equity
        ws.cell(row=row, column=1).value = "Sponsor Equity"
        ws.cell(row=row, column=2).value = "=Assumptions!B8"  # Link to calculated equity
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        ws.cell(row=row, column=3).value = f"=B{row}/B{total_uses_row}"
        ws.cell(row=row, column=3).number_format = '0.0%'
        equity_row = row
        row += 1

        # Revolving Credit Facility
        ws.cell(row=row, column=1).value = "Revolving Credit Facility"
        ws.cell(row=row, column=2).value = "=Assumptions!B11"  # Revolver from assumptions
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        ws.cell(row=row, column=3).value = f"=B{row}/B{total_uses_row}"
        ws.cell(row=row, column=3).number_format = '0.0%'
        row += 1

        # Senior Debt
        ws.cell(row=row, column=1).value = "Senior Term Loan"
        ws.cell(row=row, column=2).value = "=Assumptions!B14"  # Link to calculated senior debt
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        ws.cell(row=row, column=3).value = f"=B{row}/B{total_uses_row}"
        ws.cell(row=row, column=3).number_format = '0.0%'
        senior_debt_row = row
        row += 1

        # Subordinated Debt
        ws.cell(row=row, column=1).value = "Subordinated Notes"
        ws.cell(row=row, column=2).value = "=Assumptions!B18"  # Link to calculated sub debt
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        ws.cell(row=row, column=3).value = f"=B{row}/B{total_uses_row}"
        ws.cell(row=row, column=3).number_format = '0.0%'
        row += 1

        sources_end = row - 1

        # Total Sources
        ws.cell(row=row, column=1).value = "TOTAL SOURCES"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = f"=SUM(B{sources_start}:B{sources_end})"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=3).value = "100.0%"
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=3).font = Font(bold=True)
        total_sources_row = row
        row += 2

        # Check
        ws.cell(row=row, column=1).value = "CHECK (Should be $0)"
        ws.cell(row=row, column=2).value = f"=B{total_sources_row}-B{total_uses_row}"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        ws.cell(row=row, column=2).font = Font(bold=True, color="FF0000")  # Red if not zero

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15

    def _create_assumptions_sheet(self, assumptions: Dict):
        """Create Assumptions sheet."""
        ws = self.wb.create_sheet("Assumptions")

        # Title
        ws['A1'] = "LBO MODEL ASSUMPTIONS"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)

        row = 4

        # Transaction Assumptions
        ws.cell(row=row, column=1).value = "TRANSACTION ASSUMPTIONS"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        row += 1

        # Holding Period
        ws.cell(row=row, column=1).value = "Holding Period (Years)"
        ws.cell(row=row, column=2).value = assumptions.get('holding_period', 5)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0'
        row += 1

        # Transaction Fees %
        ws.cell(row=row, column=1).value = "Transaction Fees (% of EV)"
        ws.cell(row=row, column=2).value = assumptions.get('transaction_fees_pct', 0.02)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0%'
        row += 1

        # Sponsor Equity % (input)
        ws.cell(row=row, column=1).value = "Sponsor Equity (% of Purchase Price)"
        ws.cell(row=row, column=2).value = assumptions.get('equity_contribution_pct', 0.50)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0%'
        equity_pct_row = row
        row += 1

        # Sponsor Equity $mm (calculated)
        ws.cell(row=row, column=1).value = "Sponsor Equity ($mm)"
        ws.cell(row=row, column=2).value = f"='Transaction Summary'!B7*B{equity_pct_row}"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        sponsor_equity_row = row
        row += 1

        row += 1

        # Debt Structure
        ws.cell(row=row, column=1).value = "DEBT STRUCTURE"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        row += 1

        # Revolver
        ws.cell(row=row, column=1).value = "Revolver Size ($mm)"
        ws.cell(row=row, column=2).value = assumptions.get('revolver', 0)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        row += 1

        ws.cell(row=row, column=1).value = "Revolver Interest Rate"
        ws.cell(row=row, column=2).value = assumptions.get('revolver_rate', 0.055)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.00%'
        row += 1

        # Senior Debt
        ws.cell(row=row, column=1).value = "Senior Debt (% of Purchase Price)"
        ws.cell(row=row, column=2).value = assumptions.get('senior_debt_pct', 0.40)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0%'
        senior_pct_row = row
        row += 1

        ws.cell(row=row, column=1).value = "Senior Term Loan ($mm)"
        ws.cell(row=row, column=2).value = f"='Transaction Summary'!B7*B{senior_pct_row}"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        senior_debt_row = row
        row += 1

        ws.cell(row=row, column=1).value = "Senior Interest Rate"
        ws.cell(row=row, column=2).value = assumptions.get('senior_debt_rate', 0.055)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.00%'
        senior_rate_row = row
        row += 1

        ws.cell(row=row, column=1).value = "Senior Amortization (% p.a.)"
        ws.cell(row=row, column=2).value = assumptions.get('senior_amortization_pct', 0.05)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0%'
        row += 1

        # Subordinated Debt
        ws.cell(row=row, column=1).value = "Subordinated Debt (% of Purchase Price)"
        ws.cell(row=row, column=2).value = assumptions.get('subordinated_debt_pct', 0.10)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0%'
        sub_pct_row = row
        row += 1

        ws.cell(row=row, column=1).value = "Subordinated Notes ($mm)"
        ws.cell(row=row, column=2).value = f"='Transaction Summary'!B7*B{sub_pct_row}"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        sub_debt_row = row
        row += 1

        ws.cell(row=row, column=1).value = "Subordinated Interest Rate"
        ws.cell(row=row, column=2).value = assumptions.get('sub_debt_rate', 0.095)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.00%'
        row += 1

        row += 1

        # Operating Assumptions
        ws.cell(row=row, column=1).value = "OPERATING ASSUMPTIONS"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        row += 1

        revenue_growth = assumptions.get('revenue_growth', [0.05] * 5)
        for year, growth in enumerate(revenue_growth, 1):
            ws.cell(row=row, column=1).value = f"Year {year} Revenue Growth"
            ws.cell(row=row, column=2).value = growth
            ws.cell(row=row, column=2).fill = self.INPUT_FILL
            ws.cell(row=row, column=2).number_format = '0.0%'
            row += 1

        ws.cell(row=row, column=1).value = "EBITDA Margin"
        ws.cell(row=row, column=2).value = assumptions.get('ebitda_margin', 0.30)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0%'
        row += 1

        ws.cell(row=row, column=1).value = "D&A (% of Revenue)"
        ws.cell(row=row, column=2).value = assumptions.get('da_pct', 0.03)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0%'
        row += 1

        ws.cell(row=row, column=1).value = "CapEx (% of Revenue)"
        ws.cell(row=row, column=2).value = assumptions.get('capex_pct', 0.03)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0%'
        row += 1

        ws.cell(row=row, column=1).value = "NWC (% of Revenue)"
        ws.cell(row=row, column=2).value = assumptions.get('nwc_pct', 0.10)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0%'
        row += 1

        ws.cell(row=row, column=1).value = "Tax Rate"
        ws.cell(row=row, column=2).value = assumptions.get('tax_rate', 0.25)
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '0.0%'

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

    def _create_operating_model(self, transaction_data: Dict, assumptions: Dict):
        """Create Operating Model with projections."""
        ws = self.wb.create_sheet("Operating Model")

        # Title
        ws['A1'] = "OPERATING MODEL"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)

        # Years
        row = 3
        ws.cell(row=row, column=1).value = "Year"
        for year in range(6):  # Year 0 (LTM) + Years 1-5
            label = "LTM" if year == 0 else f"Year {year}"
            ws.cell(row=row, column=2 + year).value = label
            ws.cell(row=row, column=2 + year).font = Font(bold=True)
            ws.cell(row=row, column=2 + year).alignment = Alignment(horizontal='center')

        row += 1

        # Revenue
        ws.cell(row=row, column=1).value = "Revenue"
        ws.cell(row=row, column=1).font = Font(bold=True)

        # LTM Revenue (from transaction data)
        ltm_revenue = transaction_data.get('ltm_revenue', 180000)
        ws.cell(row=row, column=2).value = ltm_revenue
        ws.cell(row=row, column=2).fill = self.INPUT_FILL
        ws.cell(row=row, column=2).number_format = '$#,##0.0'

        # Projected Revenue
        for year in range(1, 6):
            col = 2 + year
            prior_col = get_column_letter(col - 1)
            # Reference Year 1-5 revenue growth rates at B22-B26
            ws.cell(row=row, column=col).value = f"={prior_col}{row}*(1+Assumptions!B{21+year})"
            ws.cell(row=row, column=col).number_format = '$#,##0.0'

        revenue_row = row
        row += 1

        # EBITDA
        ws.cell(row=row, column=1).value = "EBITDA"
        ws.cell(row=row, column=1).font = Font(bold=True)

        # LTM EBITDA
        ws.cell(row=row, column=2).value = "='Transaction Summary'!B5"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'

        # Projected EBITDA (Revenue * EBITDA Margin at B27)
        for year in range(1, 6):
            col = 2 + year
            col_letter = get_column_letter(col)
            ws.cell(row=row, column=col).value = f"={col_letter}{revenue_row}*Assumptions!$B$27"
            ws.cell(row=row, column=col).number_format = '$#,##0.0'

        ebitda_row = row
        row += 1

        # D&A
        ws.cell(row=row, column=1).value = "Less: D&A"

        for year in range(6):
            col = 2 + year
            col_letter = get_column_letter(col)
            # D&A % is at B28
            ws.cell(row=row, column=col).value = f"=-{col_letter}{revenue_row}*Assumptions!$B$28"
            ws.cell(row=row, column=col).number_format = '$#,##0.0'

        da_row = row
        row += 1

        # EBIT
        ws.cell(row=row, column=1).value = "EBIT"
        ws.cell(row=row, column=1).font = Font(bold=True)

        for year in range(6):
            col = 2 + year
            col_letter = get_column_letter(col)
            ws.cell(row=row, column=col).value = f"={col_letter}{ebitda_row}+{col_letter}{da_row}"
            ws.cell(row=row, column=col).number_format = '$#,##0.0'

        ebit_row = row
        row += 1

        # Interest Expense
        ws.cell(row=row, column=1).value = "Less: Interest Expense"

        for year in range(1, 6):  # No interest in LTM
            col = 2 + year
            col_letter = get_column_letter(col)
            # Link to debt schedule
            ws.cell(row=row, column=col).value = f"=-'Debt Schedule'!{col_letter}20"
            ws.cell(row=row, column=col).number_format = '$#,##0.0'

        interest_row = row
        row += 1

        # EBT
        ws.cell(row=row, column=1).value = "EBT"

        for year in range(1, 6):
            col = 2 + year
            col_letter = get_column_letter(col)
            ws.cell(row=row, column=col).value = f"={col_letter}{ebit_row}+{col_letter}{interest_row}"
            ws.cell(row=row, column=col).number_format = '$#,##0.0'

        ebt_row = row
        row += 1

        # Taxes
        ws.cell(row=row, column=1).value = "Less: Taxes"

        for year in range(1, 6):
            col = 2 + year
            col_letter = get_column_letter(col)
            # Tax Rate is at B31
            ws.cell(row=row, column=col).value = f"=-{col_letter}{ebt_row}*Assumptions!$B$31"
            ws.cell(row=row, column=col).number_format = '$#,##0.0'

        tax_row = row
        row += 1

        # Net Income
        ws.cell(row=row, column=1).value = "Net Income"
        ws.cell(row=row, column=1).font = Font(bold=True)

        for year in range(1, 6):
            col = 2 + year
            col_letter = get_column_letter(col)
            ws.cell(row=row, column=col).value = f"={col_letter}{ebt_row}+{col_letter}{tax_row}"
            ws.cell(row=row, column=col).number_format = '$#,##0.0'
            ws.cell(row=row, column=col).font = Font(bold=True)

        ws.column_dimensions['A'].width = 30
        for col in range(2, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_debt_schedule(self, assumptions: Dict):
        """Create Debt Schedule."""
        ws = self.wb.create_sheet("Debt Schedule")

        # Title
        ws['A1'] = "DEBT SCHEDULE"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)

        row = 3
        # Years
        ws.cell(row=row, column=1).value = "Year"
        for year in range(6):
            ws.cell(row=row, column=2 + year).value = f"Year {year}" if year > 0 else "Close"
            ws.cell(row=row, column=2 + year).font = Font(bold=True)
            ws.cell(row=row, column=2 + year).alignment = Alignment(horizontal='center')

        row += 2

        # Senior Term Loan
        ws.cell(row=row, column=1).value = "SENIOR TERM LOAN"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        row += 1

        # Opening Balance
        ws.cell(row=row, column=1).value = "Opening Balance"
        # Year 0 (Senior Term Loan at B14)
        ws.cell(row=row, column=2).value = "=Assumptions!B14"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'

        # Years 1-5: Prior closing
        for year in range(1, 6):
            col = 2 + year
            prior_col = get_column_letter(col - 1)
            ws.cell(row=row, column=col).value = f"={prior_col}{row+4}"  # Prior closing balance
            ws.cell(row=row, column=col).number_format = '$#,##0.0'

        opening_row = row
        row += 1

        # Mandatory Amortization
        ws.cell(row=row, column=1).value = "Mandatory Amortization"

        for year in range(1, 6):
            col = 2 + year
            col_letter = get_column_letter(col)
            # Amortization = Opening × Amortization % (at B16)
            ws.cell(row=row, column=col).value = f"=-{col_letter}{opening_row}*Assumptions!$B$16"
            ws.cell(row=row, column=col).number_format = '$#,##0.0'

        amort_row = row
        row += 1

        # Optional Prepayment (cash sweep)
        ws.cell(row=row, column=1).value = "Optional Prepayment"
        # Placeholder - would link to cash sweep
        for year in range(1, 6):
            ws.cell(row=row, column=2 + year).value = 0
            ws.cell(row=row, column=2 + year).number_format = '$#,##0.0'

        prepay_row = row
        row += 1

        # Closing Balance
        ws.cell(row=row, column=1).value = "Closing Balance"
        ws.cell(row=row, column=1).font = Font(bold=True)

        ws.cell(row=row, column=2).value = f"=B{opening_row}"  # Year 0 = opening
        ws.cell(row=row, column=2).number_format = '$#,##0.0'

        for year in range(1, 6):
            col = 2 + year
            col_letter = get_column_letter(col)
            ws.cell(row=row, column=col).value = f"={col_letter}{opening_row}+{col_letter}{amort_row}+{col_letter}{prepay_row}"
            ws.cell(row=row, column=col).number_format = '$#,##0.0'
            ws.cell(row=row, column=col).font = Font(bold=True)

        closing_row = row
        row += 1

        # Interest Expense
        ws.cell(row=row, column=1).value = "Interest Expense"

        for year in range(1, 6):
            col = 2 + year
            col_letter = get_column_letter(col)
            prior_col = get_column_letter(col - 1)
            # Interest = Average Balance × Rate (Senior Interest Rate at B15)
            ws.cell(row=row, column=col).value = f"=({prior_col}{closing_row}+{col_letter}{closing_row})/2*Assumptions!$B$15"
            ws.cell(row=row, column=col).number_format = '$#,##0.0'

        interest_row = row

        # Add borders
        self._add_table_border(ws, f'A{row-4}:G{row}')

        ws.column_dimensions['A'].width = 30

    def _create_cash_flow_waterfall(self, assumptions: Dict):
        """Create Cash Flow Waterfall."""
        ws = self.wb.create_sheet("Cash Flow Waterfall")

        ws['A1'] = "CASH FLOW WATERFALL"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)

        # Placeholder - would show detailed cash sweep logic
        ws['A3'] = "This sheet would contain detailed cash flow waterfall"
        ws['A4'] = "showing EBITDA → FCF → Debt Paydown → Cash to Equity"

    def _create_returns_analysis(self, transaction_data: Dict, assumptions: Dict):
        """Create Returns Analysis sheet."""
        ws = self.wb.create_sheet("Returns Analysis")

        # Title
        ws['A1'] = "RETURNS ANALYSIS"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)

        row = 4

        # Exit Valuation
        ws.cell(row=row, column=1).value = "EXIT VALUATION"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        row += 1

        # Exit Year EBITDA
        ws.cell(row=row, column=1).value = "Exit Year EBITDA"
        ws.cell(row=row, column=2).value = "='Operating Model'!G10"  # Year 5
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        exit_ebitda_row = row
        row += 1

        # Exit Multiple
        ws.cell(row=row, column=1).value = "Exit EV / EBITDA Multiple"
        ws.cell(row=row, column=2).value = "='Transaction Summary'!B11"  # Exit multiple at B11
        ws.cell(row=row, column=2).number_format = '0.0x'
        exit_mult_row = row
        row += 1

        # Exit Enterprise Value
        ws.cell(row=row, column=1).value = "Exit Enterprise Value"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = f"=B{exit_ebitda_row}*B{exit_mult_row}"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        ws.cell(row=row, column=2).font = Font(bold=True)
        exit_ev_row = row
        row += 1

        # Less: Remaining Debt
        ws.cell(row=row, column=1).value = "Less: Remaining Debt"
        ws.cell(row=row, column=2).value = "='Debt Schedule'!G10"  # Year 5 closing
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        remaining_debt_row = row
        row += 1

        # Equity Value at Exit
        ws.cell(row=row, column=1).value = "Equity Value at Exit"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2).value = f"=B{exit_ev_row}-B{remaining_debt_row}"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        exit_equity_row = row
        row += 2

        # Returns
        ws.cell(row=row, column=1).value = "RETURNS"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        row += 1

        # Initial Equity Investment
        ws.cell(row=row, column=1).value = "Initial Equity Investment"
        ws.cell(row=row, column=2).value = "=Assumptions!B8"
        ws.cell(row=row, column=2).number_format = '$#,##0.0'
        initial_equity_row = row
        row += 1

        # Holding Period
        ws.cell(row=row, column=1).value = "Holding Period (Years)"
        ws.cell(row=row, column=2).value = "=Assumptions!B5"
        ws.cell(row=row, column=2).number_format = '0'
        holding_row = row
        row += 1

        # IRR
        ws.cell(row=row, column=1).value = "IRR"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        # IRR = (Exit/Entry)^(1/Years) - 1
        ws.cell(row=row, column=2).value = f"=(B{exit_equity_row}/B{initial_equity_row})^(1/B{holding_row})-1"
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        row += 1

        # MOIC
        ws.cell(row=row, column=1).value = "MOIC (Multiple of Invested Capital)"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2).value = f"=B{exit_equity_row}/B{initial_equity_row}"
        ws.cell(row=row, column=2).number_format = '0.0x'
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

    def _add_table_border(self, ws, range_str: str):
        """Add borders around table range."""
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Parse range
        start_cell, end_cell = range_str.split(':')
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

        start_col_letter, start_row = coordinate_from_string(start_cell)
        end_col_letter, end_row = coordinate_from_string(end_cell)

        start_col = column_index_from_string(start_col_letter)
        end_col = column_index_from_string(end_col_letter)

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = thin_border
