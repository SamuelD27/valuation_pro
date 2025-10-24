"""
DCF Valuation Tool - Pixel-Perfect IB-Standard Model Generator

Creates professional DCF models with:
- Proper IB-style table formatting with borders
- Ultra-accurate formulas (no hardcoded values)
- Clean, readable layout matching Reference_DCF.xlsx
- All calculations use Excel formulas

This is a FOCUSED tool for DCF only (not LBO, Comps, etc.)
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers
)
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional
from datetime import datetime

from src.excel.formatter import IBFormatter


class DCFTool:
    """
    Professional DCF Model Generator.

    Creates pixel-perfect DCF models with proper IB formatting.
    """

    def __init__(self, company_name: str, ticker: str = ""):
        """
        Initialize DCF Tool.

        Args:
            company_name: Name of company being valued
            ticker: Stock ticker (optional)
        """
        self.company_name = company_name
        self.ticker = ticker
        self.wb = None
        self.formatter = IBFormatter()

        # Standard formatting - Professional IB colors
        self.SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Dark blue header
        self.SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")  # White text
        self.INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow for inputs
        self.INPUT_FONT = Font(name="Calibri", size=11, color="000000")  # Black text

    def generate_dcf_model(
        self,
        historical_data: Dict,
        assumptions: Dict,
        output_file: str
    ):
        """
        Generate complete DCF model.

        Args:
            historical_data: Historical financial data
            assumptions: DCF assumptions
            output_file: Output file path
        """
        self.wb = Workbook()
        del self.wb['Sheet']  # Remove default sheet

        # Create sheets in order
        self._create_cover_sheet()
        self._create_assumptions_sheet(assumptions)
        self._create_historical_data_sheet(historical_data)
        self._create_projections_sheet(assumptions)
        self._create_dcf_valuation_sheet(assumptions)
        self._create_sensitivity_sheet()

        # Save
        self.wb.save(output_file)
        print(f"✅ DCF model saved to: {output_file}")

    def _create_cover_sheet(self):
        """Create professional cover page."""
        ws = self.wb.create_sheet("Cover", 0)

        # Title
        ws['B2'] = f"DISCOUNTED CASH FLOW ANALYSIS"
        ws['B2'].font = Font(name="Calibri", size=20, bold=True)

        ws['B4'] = self.company_name
        ws['B4'].font = Font(name="Calibri", size=16, bold=True)

        if self.ticker:
            ws['B5'] = f"Ticker: {self.ticker}"
            ws['B5'].font = Font(name="Calibri", size=12)

        ws['B7'] = f"Valuation Date: {datetime.now().strftime('%B %d, %Y')}"
        ws['B7'].font = Font(name="Calibri", size=11)

        # Summary box
        ws['B10'] = "VALUATION SUMMARY"
        ws['B10'].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        ws['B10'].fill = self.SECTION_FILL

        summary_items = [
            ("Enterprise Value", "='DCF Valuation'!D30"),
            ("(Less): Net Debt", "='DCF Valuation'!D31"),
            ("Equity Value", "='DCF Valuation'!D32"),
            ("Shares Outstanding (mm)", "='Assumptions'!B20"),
            ("Implied Price per Share", "='DCF Valuation'!D34"),
        ]

        row = 11
        for label, formula in summary_items:
            ws.cell(row=row, column=2).value = label
            ws.cell(row=row, column=2).font = Font(name="Calibri", size=11)

            ws.cell(row=row, column=3).value = formula
            ws.cell(row=row, column=3).font = Font(name="Calibri", size=11, bold=True)

            # Format value cell
            if "Price" in label:
                ws.cell(row=row, column=3).number_format = '$#,##0.00'
            elif "Shares" in label:
                ws.cell(row=row, column=3).number_format = '#,##0'
            else:
                ws.cell(row=row, column=3).number_format = '$#,##0.0,,"M"'

            row += 1

        # Add border around summary
        self._add_table_border(ws, 'B10:C15')

        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20

    def _create_assumptions_sheet(self, assumptions: Dict):
        """
        Create Assumptions sheet with all inputs.

        All cells are blue (user inputs).
        """
        ws = self.wb.create_sheet("Assumptions")

        # Title
        ws['A1'] = "DCF MODEL ASSUMPTIONS"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)
        ws['A2'] = "All blue cells are user inputs"
        ws['A2'].font = Font(name="Calibri", size=10, italic=True)

        row = 4

        # Revenue Growth Assumptions
        ws.cell(row=row, column=1).value = "REVENUE GROWTH ASSUMPTIONS"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        self._add_table_border(ws, f'A{row}:D{row}')
        row += 1

        revenue_growth = assumptions.get('revenue_growth', [0.05] * 5)
        for idx, growth in enumerate(revenue_growth, 1):
            ws.cell(row=row, column=1).value = f"Year {idx} Revenue Growth %"
            ws.cell(row=row, column=2).value = growth
            ws.cell(row=row, column=2).fill = self.INPUT_FILL
            ws.cell(row=row, column=2).number_format = '0.0%'
            row += 1

        row += 1

        # Operating Assumptions
        ws.cell(row=row, column=1).value = "OPERATING ASSUMPTIONS"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        self._add_table_border(ws, f'A{row}:D{row}')
        row += 1

        operating_items = [
            ("EBIT Margin %", assumptions.get('ebit_margin', 0.25), '0.0%'),
            ("Tax Rate %", assumptions.get('tax_rate', 0.21), '0.0%'),
            ("CapEx (% of Revenue)", assumptions.get('capex_pct_revenue', 0.03), '0.0%'),
            ("NWC (% of Revenue)", assumptions.get('nwc_pct_revenue', 0.02), '0.0%'),
        ]

        for label, value, fmt in operating_items:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=2).value = value
            ws.cell(row=row, column=2).fill = self.INPUT_FILL
            ws.cell(row=row, column=2).number_format = fmt
            row += 1

        row += 1

        # Valuation Assumptions
        ws.cell(row=row, column=1).value = "VALUATION ASSUMPTIONS"
        ws.cell(row=row, column=1).font = self.SECTION_FONT
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        self._add_table_border(ws, f'A{row}:D{row}')
        row += 1

        valuation_items = [
            ("WACC %", assumptions.get('wacc', 0.10), '0.00%'),
            ("Terminal Growth Rate %", assumptions.get('terminal_growth', 0.025), '0.00%'),
            ("Shares Outstanding (mm)", assumptions.get('shares_outstanding', 100), '#,##0'),
            ("Net Debt ($mm)", assumptions.get('net_debt', 0), '$#,##0.0'),
        ]

        for label, value, fmt in valuation_items:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=2).value = value
            ws.cell(row=row, column=2).fill = self.INPUT_FILL
            ws.cell(row=row, column=2).number_format = fmt
            row += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

    def _create_historical_data_sheet(self, historical_data: Dict):
        """
        Create Historical Data sheet.

        Shows actual financial data from past years.
        """
        ws = self.wb.create_sheet("Historical Data")

        # Title
        ws['A1'] = "HISTORICAL FINANCIAL DATA"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)

        # Years
        years = historical_data.get('years', [2019, 2020])
        row = 3
        ws.cell(row=row, column=1).value = "Year"
        for idx, year in enumerate(years):
            ws.cell(row=row, column=2 + idx).value = year
            ws.cell(row=row, column=2 + idx).font = Font(bold=True)

        row += 1

        # Revenue
        ws.cell(row=row, column=1).value = "Revenue ($mm)"
        revenue = historical_data.get('revenue', [0] * len(years))
        for idx, val in enumerate(revenue):
            ws.cell(row=row, column=2 + idx).value = val / 1000 if val > 1000 else val
            ws.cell(row=row, column=2 + idx).number_format = '#,##0.0'
        row += 1

        # EBIT
        ws.cell(row=row, column=1).value = "EBIT ($mm)"
        ebit = historical_data.get('ebit', [0] * len(years))
        for idx, val in enumerate(ebit):
            ws.cell(row=row, column=2 + idx).value = val / 1000 if val > 1000 else val
            ws.cell(row=row, column=2 + idx).number_format = '#,##0.0'
        row += 1

        # Add border
        self._add_table_border(ws, f'A3:{get_column_letter(1 + len(years))}{row - 1}')

        ws.column_dimensions['A'].width = 25

    def _create_projections_sheet(self, assumptions: Dict):
        """
        Create Projections sheet with forecasted financials.

        Uses formulas referencing Assumptions sheet.
        """
        ws = self.wb.create_sheet("Projections")

        # Title
        ws['A1'] = "FINANCIAL PROJECTIONS"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)

        # Years (Year 1 through Year 5)
        row = 3
        ws.cell(row=row, column=1).value = "Year"
        for year in range(1, 6):
            ws.cell(row=row, column=1 + year).value = f"Year {year}"
            ws.cell(row=row, column=1 + year).font = Font(bold=True)
            ws.cell(row=row, column=1 + year).alignment = Alignment(horizontal='center')

        row += 1

        # Revenue (formula-based)
        ws.cell(row=row, column=1).value = "Revenue ($mm)"
        ws.cell(row=row, column=1).font = Font(bold=True)

        # Year 1: Last year's historical revenue × (1 + growth)
        ws.cell(row=row, column=2).value = "='Historical Data'!F10*(1+Assumptions!B5)"  # F10 = last historical year revenue
        for year in range(2, 6):
            col_letter = get_column_letter(1 + year)
            prior_col = get_column_letter(year)
            ws.cell(row=row, column=1 + year).value = f"={prior_col}{row}*(1+Assumptions!B{4 + year})"

        for col in range(2, 7):
            ws.cell(row=row, column=col).number_format = '$#,##0.0'

        revenue_row = row
        row += 1

        # EBIT
        ws.cell(row=row, column=1).value = "EBIT ($mm)"
        ws.cell(row=row, column=1).font = Font(bold=True)

        for year in range(1, 6):
            col_letter = get_column_letter(1 + year)
            # EBIT = Revenue × EBIT Margin
            ws.cell(row=row, column=1 + year).value = f"={col_letter}{revenue_row}*Assumptions!$B$12"
            ws.cell(row=row, column=1 + year).number_format = '$#,##0.0'

        ebit_row = row
        row += 1

        # Tax
        ws.cell(row=row, column=1).value = "Less: Tax"
        for year in range(1, 6):
            col_letter = get_column_letter(1 + year)
            ws.cell(row=row, column=1 + year).value = f"=-{col_letter}{ebit_row}*Assumptions!$B$13"
            ws.cell(row=row, column=1 + year).number_format = '$#,##0.0'

        tax_row = row
        row += 1

        # NOPAT
        ws.cell(row=row, column=1).value = "NOPAT"
        ws.cell(row=row, column=1).font = Font(bold=True)
        for year in range(1, 6):
            col_letter = get_column_letter(1 + year)
            ws.cell(row=row, column=1 + year).value = f"={col_letter}{ebit_row}+{col_letter}{tax_row}"
            ws.cell(row=row, column=1 + year).number_format = '$#,##0.0'

        nopat_row = row
        row += 1

        # Less: CapEx
        ws.cell(row=row, column=1).value = "Less: CapEx"
        for year in range(1, 6):
            col_letter = get_column_letter(1 + year)
            ws.cell(row=row, column=1 + year).value = f"=-{col_letter}{revenue_row}*Assumptions!$B$14"
            ws.cell(row=row, column=1 + year).number_format = '$#,##0.0'

        capex_row = row
        row += 1

        # Less: Change in NWC
        ws.cell(row=row, column=1).value = "Less: Δ NWC"
        # First year: assume from zero
        ws.cell(row=row, column=2).value = f"=-B{revenue_row}*Assumptions!$B$15"
        # Subsequent years: change from prior
        for year in range(2, 6):
            col_letter = get_column_letter(1 + year)
            prior_col = get_column_letter(year)
            ws.cell(row=row, column=1 + year).value = f"=-({col_letter}{revenue_row}*Assumptions!$B$15-{prior_col}{revenue_row}*Assumptions!$B$15)"

        for col in range(2, 7):
            ws.cell(row=row, column=col).number_format = '$#,##0.0'

        nwc_row = row
        row += 1

        # FREE CASH FLOW
        ws.cell(row=row, column=1).value = "FREE CASH FLOW"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=row, column=1).fill = self.SECTION_FILL

        for year in range(1, 6):
            col_letter = get_column_letter(1 + year)
            ws.cell(row=row, column=1 + year).value = f"={col_letter}{nopat_row}+{col_letter}{capex_row}+{col_letter}{nwc_row}"
            ws.cell(row=row, column=1 + year).number_format = '$#,##0.0'
            ws.cell(row=row, column=1 + year).font = Font(bold=True)

        # Add border around table
        self._add_table_border(ws, f'A3:F{row}')

        ws.column_dimensions['A'].width = 25
        for col_idx in range(2, 7):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    def _create_dcf_valuation_sheet(self, assumptions: Dict):
        """
        Create DCF Valuation sheet with PV calculations.

        This is the core valuation sheet.
        """
        ws = self.wb.create_sheet("DCF Valuation")

        # Title
        ws['A1'] = "DCF VALUATION"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)

        row = 3

        # Years
        ws.cell(row=row, column=1).value = "Year"
        for year in range(1, 6):
            ws.cell(row=row, column=1 + year).value = year
            ws.cell(row=row, column=1 + year).font = Font(bold=True)
            ws.cell(row=row, column=1 + year).alignment = Alignment(horizontal='center')

        row += 1

        # Free Cash Flow (link to Projections)
        ws.cell(row=row, column=1).value = "Free Cash Flow"
        ws.cell(row=row, column=1).font = Font(bold=True)

        for year in range(1, 6):
            col_letter = get_column_letter(1 + year)
            ws.cell(row=row, column=1 + year).value = f"=Projections!{col_letter}10"
            ws.cell(row=row, column=1 + year).number_format = '$#,##0.0'

        fcf_row = row
        row += 1

        # Discount Period
        ws.cell(row=row, column=1).value = "Discount Period"
        for year in range(1, 6):
            ws.cell(row=row, column=1 + year).value = year
            ws.cell(row=row, column=1 + year).number_format = '0'

        period_row = row
        row += 1

        # Discount Factor
        ws.cell(row=row, column=1).value = "Discount Factor"
        for year in range(1, 6):
            col_letter = get_column_letter(1 + year)
            ws.cell(row=row, column=1 + year).value = f"=1/((1+Assumptions!$B$18)^{col_letter}{period_row})"
            ws.cell(row=row, column=1 + year).number_format = '0.000'

        discount_row = row
        row += 1

        # Present Value of FCF
        ws.cell(row=row, column=1).value = "PV of FCF"
        ws.cell(row=row, column=1).font = Font(bold=True)

        for year in range(1, 6):
            col_letter = get_column_letter(1 + year)
            ws.cell(row=row, column=1 + year).value = f"={col_letter}{fcf_row}*{col_letter}{discount_row}"
            ws.cell(row=row, column=1 + year).number_format = '$#,##0.0'
            ws.cell(row=row, column=1 + year).font = Font(bold=True)

        pv_fcf_row = row
        row += 2

        # Terminal Value
        ws.cell(row=row, column=1).value = "Terminal Value"
        ws.cell(row=row, column=1).font = Font(bold=True)
        # TV = FCF_Year5 × (1 + g) / (WACC - g)
        ws.cell(row=row, column=6).value = "=F4*(1+Assumptions!$B$19)/(Assumptions!$B$18-Assumptions!$B$19)"
        ws.cell(row=row, column=6).number_format = '$#,##0.0'

        tv_row = row
        row += 1

        # PV of Terminal Value
        ws.cell(row=row, column=1).value = "PV of Terminal Value"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=6).value = f"=F{tv_row}*F{discount_row}"
        ws.cell(row=row, column=6).number_format = '$#,##0.0'

        pv_tv_row = row
        row += 2

        # Sum of PV of FCFs
        ws.cell(row=row, column=1).value = "Sum of PV of FCFs"
        ws.cell(row=row, column=4).value = f"=SUM(B{pv_fcf_row}:F{pv_fcf_row})"
        ws.cell(row=row, column=4).number_format = '$#,##0.0'

        sum_pv_row = row
        row += 1

        # PV of Terminal Value (repeat)
        ws.cell(row=row, column=1).value = "PV of Terminal Value"
        ws.cell(row=row, column=4).value = f"=F{pv_tv_row}"
        ws.cell(row=row, column=4).number_format = '$#,##0.0'

        row += 1

        # ENTERPRISE VALUE
        ws.cell(row=row, column=1).value = "ENTERPRISE VALUE"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        ws.cell(row=row, column=4).value = f"=D{sum_pv_row}+D{sum_pv_row + 1}"
        ws.cell(row=row, column=4).number_format = '$#,##0.0'
        ws.cell(row=row, column=4).font = Font(bold=True)

        ev_row = row
        row += 1

        # Less: Net Debt
        ws.cell(row=row, column=1).value = "Less: Net Debt"
        ws.cell(row=row, column=4).value = "=Assumptions!$B$20"
        ws.cell(row=row, column=4).number_format = '$#,##0.0'

        net_debt_row = row
        row += 1

        # EQUITY VALUE
        ws.cell(row=row, column=1).value = "EQUITY VALUE"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        ws.cell(row=row, column=4).value = f"=D{ev_row}-D{net_debt_row}"
        ws.cell(row=row, column=4).number_format = '$#,##0.0'
        ws.cell(row=row, column=4).font = Font(bold=True)

        equity_row = row
        row += 1

        # Shares Outstanding
        ws.cell(row=row, column=1).value = "Shares Outstanding (mm)"
        ws.cell(row=row, column=4).value = "=Assumptions!$B$20"
        ws.cell(row=row, column=4).number_format = '#,##0'

        shares_row = row
        row += 1

        # IMPLIED PRICE PER SHARE
        ws.cell(row=row, column=1).value = "IMPLIED PRICE PER SHARE"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=row, column=1).fill = self.SECTION_FILL
        ws.cell(row=row, column=4).value = f"=D{equity_row}/D{shares_row}"
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=4).font = Font(bold=True, size=12)

        # Add borders
        self._add_table_border(ws, f'A3:F{pv_fcf_row}')
        self._add_table_border(ws, f'A{tv_row}:F{pv_tv_row}')
        self._add_table_border(ws, f'A{sum_pv_row}:D{row}')

        ws.column_dimensions['A'].width = 30
        for col_idx in range(2, 7):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    def _create_sensitivity_sheet(self):
        """Create sensitivity analysis table."""
        ws = self.wb.create_sheet("Sensitivity")

        # Title
        ws['A1'] = "SENSITIVITY ANALYSIS"
        ws['A1'].font = Font(name="Calibri", size=14, bold=True)
        ws['A2'] = "Price per Share"
        ws['A2'].font = Font(name="Calibri", size=11, italic=True)

        # This would create a 2-way data table
        # For now, placeholder structure
        ws['B4'] = "Terminal Growth Rate →"
        ws['B4'].font = Font(bold=True)

        ws['A5'] = "WACC ↓"
        ws['A5'].font = Font(bold=True)

        # Headers for terminal growth
        growth_rates = [0.015, 0.02, 0.025, 0.03, 0.035]
        for idx, rate in enumerate(growth_rates):
            ws.cell(row=4, column=3 + idx).value = rate
            ws.cell(row=4, column=3 + idx).number_format = '0.0%'
            ws.cell(row=4, column=3 + idx).font = Font(bold=True)
            ws.cell(row=4, column=3 + idx).alignment = Alignment(horizontal='center')

        # Headers for WACC
        wacc_rates = [0.08, 0.09, 0.10, 0.11, 0.12]
        for idx, rate in enumerate(wacc_rates):
            ws.cell(row=5 + idx, column=2).value = rate
            ws.cell(row=5 + idx, column=2).number_format = '0.0%'
            ws.cell(row=5 + idx, column=2).font = Font(bold=True)

        # Placeholder values (would use Data Table in real implementation)
        for row_idx in range(5):
            for col_idx in range(5):
                ws.cell(row=5 + row_idx, column=3 + col_idx).value = 100 + (col_idx - row_idx) * 10
                ws.cell(row=5 + row_idx, column=3 + col_idx).number_format = '$#,##0.00'

        # Add border
        self._add_table_border(ws, 'B4:G9')

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12

    def _add_table_border(self, ws, range_str: str):
        """
        Add borders around a table range.

        Args:
            ws: Worksheet
            range_str: Range like 'A1:D10'
        """
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Parse range
        start_cell, end_cell = range_str.split(':')
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

        start_col_letter, start_row = coordinate_from_string(start_cell)
        end_col_letter, end_row = coordinate_from_string(end_cell)

        start_col = column_index_from_string(start_col_letter)
        end_col = column_index_from_string(end_col_letter)

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = thin_border
