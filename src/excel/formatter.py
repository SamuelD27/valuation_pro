"""
Excel Formatter - Investment Banking Standards

Provides formatting utilities to make Excel outputs look professional
and match investment banking standards (Goldman Sachs, Morgan Stanley style).

Key formatting rules:
- Blue cells = User inputs (editable)
- Black cells = Formulas (calculated)
- Thousands separators for numbers
- Proper currency and percentage formatting
- Borders around key tables
- Freeze panes for easy navigation
- Conditional formatting for negatives (red)
"""

from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NumberFormat
)
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from typing import Optional


class IBFormatter:
    """
    Investment Banking standard Excel formatting.

    Color Scheme:
    - Input cells: Light blue background (#DDEBF7), blue font (#0000FF)
    - Formula cells: White background, black font (#000000)
    - Headers: Bold, dark blue background (#4472C4), white font
    - Negative numbers: Red background (#FFC7CE), dark red font (#9C0006)
    """

    # Color definitions
    INPUT_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    INPUT_FONT = Font(color="0000FF", name="Calibri", size=11)

    FORMULA_FILL = PatternFill(fill_type=None)  # White/no fill
    FORMULA_FONT = Font(color="000000", name="Calibri", size=11)

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", name="Calibri", size=11, bold=True)

    NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    NEGATIVE_FONT = Font(color="9C0006", name="Calibri", size=11)

    TITLE_FONT = Font(name="Calibri", size=14, bold=True)

    # Border styles
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    THICK_BORDER = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    # Alignment
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center')

    @staticmethod
    def format_currency(cell, decimals: int = 0, millions: bool = True):
        """
        Format cell as currency.

        Args:
            cell: openpyxl cell object
            decimals: Number of decimal places (default 0)
            millions: If True, show in millions with 'M' suffix (default True)
        """
        if millions:
            # Format: $#,##0.0,,"M" (displays in millions)
            if decimals == 0:
                cell.number_format = '$#,##0,,"M"'
            elif decimals == 1:
                cell.number_format = '$#,##0.0,,"M"'
            else:
                cell.number_format = f'$#,##0.{"0" * decimals},,"M"'
        else:
            # Standard currency format
            if decimals == 0:
                cell.number_format = '$#,##0'
            else:
                cell.number_format = f'$#,##0.{"0" * decimals}'

    @staticmethod
    def format_percentage(cell, decimals: int = 1):
        """
        Format cell as percentage.

        Args:
            cell: openpyxl cell object
            decimals: Number of decimal places (default 1)
        """
        if decimals == 0:
            cell.number_format = '0%'
        elif decimals == 1:
            cell.number_format = '0.0%'
        else:
            cell.number_format = f'0.{"0" * decimals}%'

    @staticmethod
    def format_number(cell, decimals: int = 1, thousands_sep: bool = True):
        """
        Format cell as number with thousands separator.

        Args:
            cell: openpyxl cell object
            decimals: Number of decimal places (default 1)
            thousands_sep: Include thousands separator (default True)
        """
        if thousands_sep:
            if decimals == 0:
                cell.number_format = '#,##0'
            else:
                cell.number_format = f'#,##0.{"0" * decimals}'
        else:
            if decimals == 0:
                cell.number_format = '0'
            else:
                cell.number_format = f'0.{"0" * decimals}'

    @staticmethod
    def format_multiple(cell, decimals: int = 1):
        """
        Format cell as multiple (e.g., 5.2x).

        Args:
            cell: openpyxl cell object
            decimals: Number of decimal places (default 1)
        """
        if decimals == 1:
            cell.number_format = '0.0"x"'
        else:
            cell.number_format = f'0.{"0" * decimals}"x"'

    @staticmethod
    def apply_input_style(cell):
        """Apply input cell style (blue background, blue font)."""
        cell.fill = IBFormatter.INPUT_FILL
        cell.font = IBFormatter.INPUT_FONT

    @staticmethod
    def apply_formula_style(cell):
        """Apply formula cell style (white background, black font)."""
        cell.fill = IBFormatter.FORMULA_FILL
        cell.font = IBFormatter.FORMULA_FONT

    @staticmethod
    def apply_header_style(cell):
        """Apply header cell style (dark blue background, white bold font)."""
        cell.fill = IBFormatter.HEADER_FILL
        cell.font = IBFormatter.HEADER_FONT
        cell.alignment = IBFormatter.CENTER_ALIGN

    @staticmethod
    def apply_title_style(cell):
        """Apply title style (large bold font)."""
        cell.font = IBFormatter.TITLE_FONT

    @staticmethod
    def apply_borders(ws: Worksheet, cell_range: str, border_style: str = "thin"):
        """
        Apply borders around a cell range.

        Args:
            ws: Worksheet object
            cell_range: Cell range (e.g., 'A1:D10')
            border_style: 'thin' or 'thick' (default 'thin')
        """
        border = IBFormatter.THIN_BORDER if border_style == "thin" else IBFormatter.THICK_BORDER

        # Parse cell range
        start_cell, end_cell = cell_range.split(':')

        # Get row and column indices
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

        start_col, start_row = coordinate_from_string(start_cell)
        end_col, end_row = coordinate_from_string(end_cell)

        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)

        # Apply borders to all cells in range
        for row in range(start_row, end_row + 1):
            for col in range(start_col_idx, end_col_idx + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border

    @staticmethod
    def set_column_width(ws: Worksheet, column: str, width: float):
        """
        Set column width.

        Args:
            ws: Worksheet object
            column: Column letter (e.g., 'A')
            width: Width in character units (default Excel is 8.43)
        """
        ws.column_dimensions[column].width = width

    @staticmethod
    def auto_size_columns(ws: Worksheet, min_width: float = 10, max_width: float = 50):
        """
        Auto-size all columns based on content.

        Args:
            ws: Worksheet object
            min_width: Minimum column width (default 10)
            max_width: Maximum column width (default 50)
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def freeze_panes(ws: Worksheet, row: int = 1, col: int = 1):
        """
        Freeze panes at specified row/column.

        Args:
            ws: Worksheet object
            row: Row to freeze above (1-indexed)
            col: Column to freeze left of (1-indexed)
        """
        # Freeze panes at cell (row+1, col+1)
        freeze_cell = f"{get_column_letter(col + 1)}{row + 1}"
        ws.freeze_panes = freeze_cell

    @staticmethod
    def add_conditional_formatting_negatives(ws: Worksheet, cell_range: str):
        """
        Add conditional formatting to highlight negative numbers in red.

        Args:
            ws: Worksheet object
            cell_range: Cell range to apply formatting (e.g., 'B5:D20')
        """
        from openpyxl.formatting.rule import CellIsRule

        red_fill = IBFormatter.NEGATIVE_FILL
        red_font = IBFormatter.NEGATIVE_FONT

        rule = CellIsRule(
            operator='lessThan',
            formula=['0'],
            fill=red_fill,
            font=red_font
        )

        ws.conditional_formatting.add(cell_range, rule)

    @staticmethod
    def create_table_header(
        ws: Worksheet,
        start_row: int,
        start_col: int,
        headers: list
    ):
        """
        Create formatted table header row.

        Args:
            ws: Worksheet object
            start_row: Starting row (1-indexed)
            start_col: Starting column (1-indexed)
            headers: List of header strings
        """
        for idx, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col + idx)
            cell.value = header
            IBFormatter.apply_header_style(cell)

    @staticmethod
    def format_sensitivity_table(
        ws: Worksheet,
        start_row: int,
        start_col: int,
        row_headers: list,
        col_headers: list,
        format_type: str = "currency"
    ):
        """
        Format a 2-way sensitivity table.

        Args:
            ws: Worksheet object
            start_row: Starting row for table (1-indexed)
            start_col: Starting column for table (1-indexed)
            row_headers: List of row header values (e.g., WACC values)
            col_headers: List of column header values (e.g., terminal growth)
            format_type: 'currency', 'percentage', or 'number'
        """
        # Top-left corner cell (empty or label)
        corner_cell = ws.cell(row=start_row, column=start_col)
        corner_cell.fill = IBFormatter.HEADER_FILL

        # Column headers
        for idx, header in enumerate(col_headers):
            cell = ws.cell(row=start_row, column=start_col + idx + 1)
            cell.value = header
            IBFormatter.apply_header_style(cell)

        # Row headers
        for idx, header in enumerate(row_headers):
            cell = ws.cell(row=start_row + idx + 1, column=start_col)
            cell.value = header
            IBFormatter.apply_header_style(cell)

        # Format data cells
        for row_idx in range(len(row_headers)):
            for col_idx in range(len(col_headers)):
                cell = ws.cell(
                    row=start_row + row_idx + 1,
                    column=start_col + col_idx + 1
                )

                # Apply appropriate number format
                if format_type == "currency":
                    IBFormatter.format_currency(cell, decimals=2, millions=False)
                elif format_type == "percentage":
                    IBFormatter.format_percentage(cell, decimals=1)
                else:
                    IBFormatter.format_number(cell, decimals=2)

                # Formula cells are black
                IBFormatter.apply_formula_style(cell)

        # Apply borders around entire table
        end_row = start_row + len(row_headers)
        end_col = start_col + len(col_headers)
        table_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        IBFormatter.apply_borders(ws, table_range, border_style="thin")
