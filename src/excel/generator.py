"""
Excel Generator - Create IB-Quality Valuation Models

Generates Excel workbooks for different valuation models:
- DCF (Discounted Cash Flow)
- LBO (Leveraged Buyout)
- Comparable Companies
- Precedent Transactions
- Merger Models

All outputs follow investment banking formatting standards.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from typing import Dict, List, Optional
from datetime import datetime

from src.excel.formatter import IBFormatter


class ExcelGenerator:
    """
    Generate investment banking-quality Excel valuation models.

    Creates formatted workbooks with:
    - Assumptions sheet (blue inputs)
    - Data sheets (historical financials)
    - Calculation sheets (formulas, not hardcoded values)
    - Valuation summary
    - Sensitivity analysis tables
    """

    def __init__(self, model_type: str):
        """
        Initialize Excel Generator.

        Args:
            model_type: Type of model - 'dcf', 'lbo', 'comps', 'precedents', 'merger'
        """
        valid_types = ['dcf', 'lbo', 'comps', 'precedents', 'merger']
        if model_type.lower() not in valid_types:
            raise ValueError(
                f"Invalid model type: {model_type}. "
                f"Must be one of {valid_types}"
            )

        self.model_type = model_type.lower()
        self.wb = None
        self.formatter = IBFormatter()

    def create_workbook(self) -> Workbook:
        """
        Initialize new workbook with IB-standard formatting.

        Returns:
            openpyxl Workbook object
        """
        self.wb = Workbook()

        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        return self.wb

    def add_assumptions_sheet(self, assumptions: Dict, sheet_name: str = "Assumptions"):
        """
        Add assumptions sheet with all user inputs.

        All cells in this sheet are formatted as inputs (blue).

        Args:
            assumptions: Dictionary of assumption categories and values
            sheet_name: Name of sheet (default 'Assumptions')
        """
        if self.wb is None:
            self.create_workbook()

        ws = self.wb.create_sheet(sheet_name, 0)  # Insert as first sheet

        # Title
        ws['A1'] = f"{self.model_type.upper()} Model - Assumptions"
        self.formatter.apply_title_style(ws['A1'])
        ws['A2'] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"

        current_row = 4

        # Iterate through assumption categories
        for category, values in assumptions.items():
            # Category header
            ws.cell(row=current_row, column=1).value = category.replace('_', ' ').title()
            self.formatter.apply_header_style(ws.cell(row=current_row, column=1))
            current_row += 1

            # Add values
            if isinstance(values, dict):
                for key, value in values.items():
                    # Label
                    label_cell = ws.cell(row=current_row, column=1)
                    label_cell.value = key.replace('_', ' ').title()

                    # Value
                    value_cell = ws.cell(row=current_row, column=2)
                    value_cell.value = value
                    self.formatter.apply_input_style(value_cell)

                    # Format based on value type
                    if isinstance(value, float):
                        if 'rate' in key.lower() or 'growth' in key.lower() or \
                           'margin' in key.lower() or 'pct' in key.lower():
                            self.formatter.format_percentage(value_cell, decimals=1)
                        elif 'price' in key.lower() or 'value' in key.lower():
                            self.formatter.format_currency(value_cell, decimals=2, millions=False)
                        else:
                            self.formatter.format_number(value_cell, decimals=2)

                    current_row += 1

            elif isinstance(values, list):
                # List of values (e.g., revenue growth rates)
                for idx, value in enumerate(values):
                    label_cell = ws.cell(row=current_row, column=1)
                    label_cell.value = f"Year {idx + 1}"

                    value_cell = ws.cell(row=current_row, column=2)
                    value_cell.value = value
                    self.formatter.apply_input_style(value_cell)

                    if isinstance(value, float):
                        self.formatter.format_percentage(value_cell, decimals=1)

                    current_row += 1

            else:
                # Single value
                value_cell = ws.cell(row=current_row, column=2)
                value_cell.value = values
                self.formatter.apply_input_style(value_cell)
                current_row += 1

            current_row += 1  # Space between categories

        # Formatting
        self.formatter.set_column_width(ws, 'A', 30)
        self.formatter.set_column_width(ws, 'B', 20)
        self.formatter.freeze_panes(ws, row=3, col=0)

    def add_data_sheet(
        self,
        data: pd.DataFrame,
        sheet_name: str,
        title: Optional[str] = None
    ):
        """
        Add data sheet from DataFrame with formatting.

        Args:
            data: DataFrame to write
            sheet_name: Name of sheet
            title: Optional title for sheet (displayed in A1)
        """
        if self.wb is None:
            self.create_workbook()

        ws = self.wb.create_sheet(sheet_name)

        start_row = 1

        # Add title if provided
        if title:
            ws['A1'] = title
            self.formatter.apply_title_style(ws['A1'])
            start_row = 3

        # Write DataFrame headers
        for col_idx, column in enumerate(data.columns, start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = column
            self.formatter.apply_header_style(cell)

        # Write DataFrame data
        for row_idx, row in enumerate(data.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                self.formatter.apply_formula_style(cell)

                # Format numbers
                if isinstance(value, (int, float)) and col_idx > 1:  # Skip first column (labels)
                    # Determine format type from column name
                    col_name = data.columns[col_idx - 1].lower()

                    if 'growth' in col_name or 'margin' in col_name or \
                       'rate' in col_name or '%' in col_name:
                        self.formatter.format_percentage(cell, decimals=1)
                    elif 'fcf' in col_name or 'revenue' in col_name or \
                         'ebit' in col_name or 'capex' in col_name or \
                         'nwc' in col_name or 'value' in col_name:
                        self.formatter.format_currency(cell, decimals=1, millions=True)
                    else:
                        self.formatter.format_number(cell, decimals=1)

        # Auto-size columns
        self.formatter.auto_size_columns(ws)

        # Freeze header row
        self.formatter.freeze_panes(ws, row=start_row, col=0)

    def add_formulas_sheet(
        self,
        sheet_name: str,
        formula_dict: Dict,
        title: Optional[str] = None
    ):
        """
        Add sheet with Excel formulas (not Python calculations).

        Args:
            sheet_name: Name of sheet
            formula_dict: Dict mapping cell addresses to formulas
                         e.g., {'B5': '=B3*B4', 'C5': '=SUM(B5:B10)'}
            title: Optional title for sheet
        """
        if self.wb is None:
            self.create_workbook()

        ws = self.wb.create_sheet(sheet_name)

        # Add title if provided
        if title:
            ws['A1'] = title
            self.formatter.apply_title_style(ws['A1'])

        # Write formulas
        for cell_address, formula in formula_dict.items():
            cell = ws[cell_address]

            # Formulas should start with '='
            if not formula.startswith('='):
                formula = '=' + formula

            cell.value = formula
            self.formatter.apply_formula_style(cell)

    def add_sensitivity_table(
        self,
        ws,
        start_row: int,
        start_col: int,
        sensitivity_df: pd.DataFrame,
        row_label: str = "WACC",
        col_label: str = "Terminal Growth",
        title: Optional[str] = None
    ):
        """
        Create Excel sensitivity table (2-way data table).

        Args:
            ws: Worksheet object
            start_row: Starting row for table (1-indexed)
            start_col: Starting column for table (1-indexed)
            sensitivity_df: DataFrame with sensitivity results
            row_label: Label for row variable
            col_label: Label for column variable
            title: Optional title above table
        """
        # Add title if provided
        if title:
            title_cell = ws.cell(row=start_row, column=start_col)
            title_cell.value = title
            self.formatter.apply_title_style(title_cell)
            start_row += 2

        # Top-left corner label
        corner_cell = ws.cell(row=start_row, column=start_col)
        corner_cell.value = f"{row_label} / {col_label}"
        self.formatter.apply_header_style(corner_cell)

        # Column headers (terminal growth rates)
        for col_idx, col_header in enumerate(sensitivity_df.columns, start=1):
            cell = ws.cell(row=start_row, column=start_col + col_idx)
            cell.value = col_header
            self.formatter.apply_header_style(cell)

        # Row headers (WACC values) and data
        for row_idx, (row_header, row_data) in enumerate(sensitivity_df.iterrows(), start=1):
            # Row header
            header_cell = ws.cell(row=start_row + row_idx, column=start_col)
            header_cell.value = row_header
            self.formatter.apply_header_style(header_cell)

            # Data cells
            for col_idx, value in enumerate(row_data, start=1):
                data_cell = ws.cell(row=start_row + row_idx, column=start_col + col_idx)
                data_cell.value = value
                self.formatter.apply_formula_style(data_cell)
                self.formatter.format_currency(data_cell, decimals=2, millions=False)

        # Apply borders
        end_row = start_row + len(sensitivity_df)
        end_col = start_col + len(sensitivity_df.columns)
        table_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        self.formatter.apply_borders(ws, table_range, border_style="thin")

        # Highlight base case (middle cell) with thick border
        mid_row = start_row + (len(sensitivity_df) // 2) + 1
        mid_col = start_col + (len(sensitivity_df.columns) // 2) + 1
        base_case_cell = ws.cell(row=mid_row, column=mid_col)
        base_case_cell.border = self.formatter.THICK_BORDER

    def apply_ib_formatting(self):
        """
        Apply investment banking formatting standards to entire workbook.

        - Freeze panes on all sheets
        - Auto-size columns
        - Apply borders to tables
        """
        if self.wb is None:
            return

        for ws in self.wb.worksheets:
            # Auto-size columns
            self.formatter.auto_size_columns(ws)

            # Freeze top row if not already frozen
            if not ws.freeze_panes:
                self.formatter.freeze_panes(ws, row=1, col=0)

    def save(self, filepath: str):
        """
        Save workbook to file.

        Args:
            filepath: Path to save Excel file (should end in .xlsx)
        """
        if self.wb is None:
            raise ValueError("No workbook created. Call create_workbook() first.")

        # Ensure filepath ends with .xlsx
        if not filepath.endswith('.xlsx'):
            filepath += '.xlsx'

        # Apply final formatting
        self.apply_ib_formatting()

        # Save
        self.wb.save(filepath)
        print(f"Excel file saved to: {filepath}")

    def create_dcf_excel(
        self,
        dcf_model,
        assumptions: Dict,
        company_data: Dict,
        filepath: str
    ):
        """
        Create complete DCF Excel model.

        Args:
            dcf_model: DCFModel instance with calculations completed
            assumptions: Assumptions dict
            company_data: Historical company data
            filepath: Path to save file
        """
        self.create_workbook()

        # 1. Summary Sheet
        self._create_dcf_summary_sheet(dcf_model)

        # 2. Assumptions Sheet
        self.add_assumptions_sheet(assumptions, sheet_name="Assumptions")

        # 3. Historical Data Sheet
        self._create_historical_sheet(company_data)

        # 4. Projections Sheet
        if dcf_model.projections is not None:
            self.add_data_sheet(
                dcf_model.projections,
                sheet_name="Projections",
                title="Financial Projections"
            )

        # 5. Valuation Sheet with Sensitivity
        self._create_dcf_valuation_sheet(dcf_model)

        # Save
        self.save(filepath)

    def _create_dcf_summary_sheet(self, dcf_model):
        """Create DCF summary sheet with key metrics."""
        ws = self.wb.create_sheet("Summary", 0)

        ws['A1'] = "DCF Valuation Summary"
        self.formatter.apply_title_style(ws['A1'])

        ws['A2'] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"

        # Valuation results
        result = dcf_model.equity_value_result

        row = 4
        ws.cell(row=row, column=1).value = "Valuation Results"
        self.formatter.apply_header_style(ws.cell(row=row, column=1))
        row += 1

        metrics = [
            ("Enterprise Value", result['enterprise_value']),
            ("Less: Net Debt", result['net_debt']),
            ("Equity Value", result['equity_value']),
            ("Shares Outstanding", result['shares_outstanding']),
            ("Price per Share", result['price_per_share']),
        ]

        if 'current_price' in result:
            metrics.append(("Current Price", result['current_price']))
            metrics.append(("Upside/(Downside)", result['upside_downside_pct']))

        for label, value in metrics:
            ws.cell(row=row, column=1).value = label
            cell = ws.cell(row=row, column=2)
            cell.value = value
            self.formatter.apply_formula_style(cell)

            if 'price' in label.lower() and 'share' not in label.lower():
                self.formatter.format_currency(cell, decimals=0, millions=True)
            elif 'share' in label.lower():
                self.formatter.format_currency(cell, decimals=2, millions=False)
            elif 'upside' in label.lower():
                self.formatter.format_percentage(cell, decimals=1)
            elif 'shares' in label.lower():
                self.formatter.format_number(cell, decimals=0)
            else:
                self.formatter.format_currency(cell, decimals=0, millions=True)

            row += 1

        self.formatter.set_column_width(ws, 'A', 30)
        self.formatter.set_column_width(ws, 'B', 20)

    def _create_historical_sheet(self, company_data: Dict):
        """Create sheet with historical financial data."""
        # This would parse company_data and create a formatted historical financials sheet
        # For now, just create a placeholder
        ws = self.wb.create_sheet("Historical")

        ws['A1'] = "Historical Financials"
        self.formatter.apply_title_style(ws['A1'])

    def _create_dcf_valuation_sheet(self, dcf_model):
        """Create valuation sheet with NPV calculation and sensitivity."""
        ws = self.wb.create_sheet("Valuation")

        ws['A1'] = "DCF Valuation & Sensitivity Analysis"
        self.formatter.apply_title_style(ws['A1'])

        # Add sensitivity table
        sensitivity_df = dcf_model.sensitivity_analysis()

        self.add_sensitivity_table(
            ws,
            start_row=5,
            start_col=1,
            sensitivity_df=sensitivity_df,
            row_label="WACC",
            col_label="Terminal Growth",
            title="Price per Share Sensitivity"
        )

        self.formatter.set_column_width(ws, 'A', 15)
