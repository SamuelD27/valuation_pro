"""
Intelligent Excel data extractor for ValuationPro.

This extractor can handle Excel files with unknown layouts, using:
- Automatic sheet detection (finds financial statements)
- Smart table detection (finds data tables in sheets)
- Fuzzy field matching (handles variations in naming)
- Layout inference (row-wise vs column-wise)
- Unit detection and conversion

No hard-coded cell references needed!
"""

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from typing import List, Tuple, Optional, Dict, Any
from pathlib import Path
import re
from rapidfuzz import fuzz, process

from .base_extractor import BaseExtractor
from ..schema import (
    FinancialData,
    CompanyInfo,
    IncomeStatement,
    BalanceSheet,
    CashFlowStatement,
    MarketData,
    ExtractionMetadata,
)
from datetime import datetime


class ExcelExtractor(BaseExtractor):
    """
    Intelligent Excel file parser.

    Can handle various Excel layouts without hardcoded cell references.
    Uses fuzzy matching and pattern detection to find financial data.
    """

    # Field synonym dictionaries for fuzzy matching
    FIELD_SYNONYMS = {
        # Income statement
        'revenue': ['sales', 'net sales', 'total revenue', 'revenues', 'turnover', 'net revenue'],
        'cogs': ['cost of goods sold', 'cost of sales', 'cost of revenue', 'direct costs'],
        'gross_profit': ['gross profit', 'gross margin', 'gross income'],
        'operating_expenses': ['operating expenses', 'opex', 'sg&a', 'sga'],
        'rd_expense': ['r&d', 'research and development', 'r & d', 'rd expense'],
        'sga_expense': ['sg&a', 'sga', 'selling general administrative', 'selling, general and administrative'],
        'ebitda': ['ebitda', 'operating income before d&a', 'adjusted ebitda', 'ebitda (adjusted)'],
        'depreciation_amortization': ['depreciation & amortization', 'd&a', 'da', 'depreciation and amortization', 'dep & amort'],
        'ebit': ['ebit', 'operating income', 'operating profit', 'oper income'],
        'interest_expense': ['interest expense', 'interest', 'interest paid', 'net interest'],
        'interest_income': ['interest income', 'interest revenue'],
        'pretax_income': ['pre-tax income', 'pretax income', 'income before tax', 'ebt'],
        'income_tax': ['income tax', 'tax', 'taxes', 'provision for taxes', 'tax expense'],
        'net_income': ['net income', 'net profit', 'profit', 'earnings', 'net earnings'],

        # Balance sheet
        'cash': ['cash', 'cash and equivalents', 'cash & cash equivalents', 'cash and short term investments'],
        'marketable_securities': ['marketable securities', 'short-term investments', 'short term investments'],
        'accounts_receivable': ['accounts receivable', 'receivables', 'trade receivables', 'a/r', 'ar'],
        'inventory': ['inventory', 'inventories', 'stock'],
        'current_assets': ['current assets', 'total current assets'],
        'ppe_net': ['property plant equipment', 'ppe', 'pp&e', 'net ppe', 'fixed assets', 'property, plant and equipment'],
        'ppe_gross': ['gross ppe', 'pp&e gross', 'property plant equipment gross'],
        'goodwill': ['goodwill'],
        'intangible_assets': ['intangible assets', 'intangibles'],
        'total_assets': ['total assets', 'assets'],

        'accounts_payable': ['accounts payable', 'payables', 'trade payables', 'a/p', 'ap'],
        'accrued_expenses': ['accrued expenses', 'accruals', 'accrued liabilities'],
        'short_term_debt': ['short-term debt', 'short term debt', 'current debt'],
        'current_liabilities': ['current liabilities', 'total current liabilities'],
        'long_term_debt': ['long-term debt', 'long term debt', 'lt debt', 'non-current debt'],
        'total_liabilities': ['total liabilities', 'liabilities'],
        'shareholders_equity': ['shareholders equity', 'shareholder equity', "shareholders' equity", 'equity', 'total equity', 'stockholders equity'],

        # Cash flow statement
        'operating_cash_flow': ['cash from operations', 'operating cash flow', 'cfo', 'cash flow from operating activities'],
        'capex': ['capex', 'capital expenditures', 'capital expenditure', 'purchase of ppe', 'additions to ppe'],
        'change_in_nwc': ['change in nwc', 'change in working capital', 'working capital change', 'delta nwc'],
        'free_cash_flow': ['free cash flow', 'fcf', 'unlevered free cash flow'],
    }

    def can_handle(self, source: Any) -> bool:
        """
        Check if source is an Excel file (.xlsx or .xls).

        Args:
            source: Input source (file path)

        Returns:
            True if Excel file, False otherwise
        """
        if isinstance(source, str):
            path = Path(source)
            return path.suffix.lower() in ['.xlsx', '.xls', '.xlsm']
        return False

    def extract(self, source: str, **kwargs) -> FinancialData:
        """
        Smart extraction from Excel file.

        Args:
            source: Path to Excel file
            **kwargs: Optional hints
                - company_name: Override detected company name
                - sheet_name: Force specific sheet
                - year_range: Tuple (start_year, end_year)

        Returns:
            FinancialData object

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If no financial data found
        """
        # Validate file exists
        path = Path(source)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {source}")

        print(f"📊 Opening Excel file: {path.name}")

        # Load workbook
        try:
            wb = openpyxl.load_workbook(source, data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to open Excel file: {e}")

        try:
            # Step 1: Detect company name
            company_name = kwargs.get('company_name') or self._detect_company_name(wb)
            print(f"✓ Company: {company_name}")

            # Step 2: Find financial statement sheets
            if 'sheet_name' in kwargs:
                sheet_names = [kwargs['sheet_name']]
            else:
                sheet_names = self._find_financial_sheets(wb)
            print(f"✓ Found {len(sheet_names)} sheet(s): {', '.join(sheet_names)}")

            # Step 3: Extract data from sheets
            all_data = {}
            for sheet_name in sheet_names:
                sheet_data = self._extract_from_sheet(wb[sheet_name])
                all_data[sheet_name] = sheet_data

            # Step 4: Combine data from multiple sheets
            combined = self._combine_sheet_data(all_data)

            # Step 5: Build FinancialData object
            financial_data = self._build_financial_data(
                company_name=company_name,
                combined_data=combined,
                source=source
            )

            print(f"✓ Extraction complete: {len(financial_data.years)} years")

        finally:
            wb.close()

        return financial_data

    def _detect_company_name(self, wb) -> str:
        """
        Try to detect company name from Excel file.

        Looks in first sheet, first few rows for company name.
        """
        first_sheet = wb.worksheets[0]

        # Check first 10 rows for company name patterns
        for row in range(1, min(11, first_sheet.max_row + 1)):
            for col in range(1, min(6, first_sheet.max_column + 1)):
                cell_value = first_sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    cell_value = cell_value.strip()

                    # Look for keywords
                    keywords = ['inc', 'corp', 'company', 'ltd', 'llc', 'plc']
                    if any(kw in cell_value.lower() for kw in keywords):
                        # Likely a company name
                        return cell_value

        # Fallback: use filename
        return Path(wb.path).stem if hasattr(wb, 'path') else "Unknown Company"

    def _find_financial_sheets(self, wb) -> List[str]:
        """
        Identify which sheets contain financial data.

        Args:
            wb: Workbook object

        Returns:
            List of sheet names likely containing financials
        """
        keywords = ['income', 'profit', 'loss', 'p&l', 'pnl', 'balance', 'cash flow', 'cashflow', 'cf', 'financial', 'statements']
        exclude = ['cover', 'summary', 'assumptions', 'model', 'notes', 'charts', 'graphs']

        candidate_sheets = []

        for sheet in wb.worksheets:
            sheet_name_lower = sheet.title.lower()

            # Skip if contains exclude keywords
            if any(excl in sheet_name_lower for excl in exclude):
                continue

            # Include if contains financial keywords
            if any(kw in sheet_name_lower for kw in keywords):
                candidate_sheets.append(sheet.title)
                continue

            # Also check if sheet has data that looks financial
            if self._sheet_has_financial_data(sheet):
                candidate_sheets.append(sheet.title)

        # If no candidates found, use first sheet as fallback
        if not candidate_sheets and len(wb.worksheets) > 0:
            candidate_sheets = [wb.worksheets[0].title]

        return candidate_sheets

    def _sheet_has_financial_data(self, sheet: Worksheet) -> bool:
        """
        Check if sheet contains financial data patterns.

        Looks for years, monetary values, and financial keywords.
        """
        years_found = 0
        financial_keywords_found = 0

        for row in sheet.iter_rows(max_row=50, max_col=20):
            for cell in row:
                value = cell.value

                # Check for years (1990-2050)
                if isinstance(value, (int, float)) and 1990 <= value <= 2050:
                    years_found += 1

                # Check for financial keywords
                if isinstance(value, str):
                    value_lower = value.lower()
                    if any(kw in value_lower for kw in ['revenue', 'income', 'ebitda', 'assets', 'liabilities', 'cash flow']):
                        financial_keywords_found += 1

        # Sheet likely has financial data if it has years and keywords
        return years_found >= 2 and financial_keywords_found >= 3

    def _extract_from_sheet(self, sheet: Worksheet) -> Dict[str, Any]:
        """
        Extract financial data from a single sheet.

        Returns:
            Dict with years and extracted metrics
        """
        # Convert sheet to DataFrame for easier manipulation
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        df = pd.DataFrame(data)

        # Find years
        years, years_location = self._find_years(df)
        if not years:
            return {'years': [], 'metrics': {}}

        print(f"  ✓ Found years: {years} ({years_location})")

        # Extract metrics based on layout
        if years_location == 'row':
            # Years in a row -> metrics in rows (NOT columns!)
            # When years are [2021, 2022, 2023...] in a row,
            # the data rows below them have [1200, 1350, 1520...]
            metrics = self._extract_rowwise(df, years)
        else:
            # Years in a column -> metrics in columns
            metrics = self._extract_columnwise(df, years)

        print(f"  ✓ Extracted {len(metrics)} metrics")

        return {'years': years, 'metrics': metrics}

    def _find_years(self, df: pd.DataFrame) -> Tuple[List[int], str]:
        """
        Find year values in the DataFrame.

        Returns:
            (list of years, location: 'row' or 'column')
        """
        years = []
        location = 'row'

        # Search first 100 rows and 30 columns for years
        # Increased from 20 to handle files with assumptions at top
        max_search_rows = min(100, len(df))
        max_search_cols = min(30, len(df.columns))

        # Search rows for years
        for row_idx in range(max_search_rows):
            row_years = []
            for col_idx in range(max_search_cols):
                value = df.iloc[row_idx, col_idx]
                if isinstance(value, (int, float)) and 1990 <= value <= 2050:
                    row_years.append(int(value))

            # If we found 2+ years in this row (could be consecutive or not)
            if len(row_years) >= 2:
                # Check if years are reasonably sequential
                sorted_years = sorted(row_years)
                year_gaps = [sorted_years[i+1] - sorted_years[i] for i in range(len(sorted_years)-1)]

                # Accept if most gaps are 1 year (consecutive) or all gaps are same (e.g., every 2 years)
                if all(gap <= 5 for gap in year_gaps):  # Reasonable gaps
                    years = row_years
                    location = 'row'
                    break

        # If not found in rows, search columns
        if not years:
            for col_idx in range(max_search_cols):
                col_years = []
                for row_idx in range(max_search_rows):
                    value = df.iloc[row_idx, col_idx]
                    if isinstance(value, (int, float)) and 1990 <= value <= 2050:
                        col_years.append(int(value))

                if len(col_years) >= 2:
                    # Check if years are reasonably sequential
                    sorted_years = sorted(col_years)
                    year_gaps = [sorted_years[i+1] - sorted_years[i] for i in range(len(sorted_years)-1)]

                    if all(gap <= 5 for gap in year_gaps):
                        years = col_years
                        location = 'column'
                        break

        return years, location

    def _extract_rowwise(self, df: pd.DataFrame, years: List[int]) -> Dict[str, List[float]]:
        """
        Extract metrics when layout is row-wise (metrics in rows, years in columns).

        Args:
            df: DataFrame of sheet data
            years: List of years (in column headers)

        Returns:
            Dict mapping standard field names to values
        """
        metrics = {}

        # Find year column indices (search more rows now - up to 100)
        year_cols = []
        for col_idx in range(len(df.columns)):
            for row_idx in range(min(100, len(df))):
                value = df.iloc[row_idx, col_idx]
                if value in years:
                    year_cols.append((col_idx, years.index(value)))
                    break

        if not year_cols:
            return metrics

        # Sort by year order
        year_cols = sorted(year_cols, key=lambda x: x[1])
        year_indices = [col for col, _ in year_cols]

        # Scan rows for metric labels
        for row_idx in range(len(df)):
            # Get row label (usually in first column)
            label = df.iloc[row_idx, 0]
            if not isinstance(label, str):
                continue

            # Clean label - remove units like ($mm), ($M), etc.
            label_clean = re.sub(r'\s*\([^)]*\)\s*', ' ', label).strip()

            # Try to match label to standard field
            standard_field = self._fuzzy_match_field(label_clean)
            if not standard_field:
                continue

            # Extract values for each year
            values = []
            for col_idx in year_indices:
                value = df.iloc[row_idx, col_idx]
                if isinstance(value, (int, float)):
                    values.append(float(value))
                else:
                    values.append(None)

            # Only add if we got at least some non-None values
            if any(v is not None for v in values):
                metrics[standard_field] = values

        return metrics

    def _extract_columnwise(self, df: pd.DataFrame, years: List[int]) -> Dict[str, List[float]]:
        """
        Extract metrics when layout is column-wise (metrics in columns, years in rows).

        Args:
            df: DataFrame of sheet data
            years: List of years (in rows)

        Returns:
            Dict mapping standard field names to values
        """
        metrics = {}

        # Find year row indices
        year_rows = []
        for row_idx in range(min(20, len(df))):
            for col_idx in range(min(20, len(df.columns))):
                value = df.iloc[row_idx, col_idx]
                if value in years:
                    year_rows.append((row_idx, years.index(value)))
                    break

        if not year_rows:
            return metrics

        # Sort by year order
        year_rows = sorted(year_rows, key=lambda x: x[1])
        year_row_indices = [row for row, _ in year_rows]

        # Scan columns for metric labels
        for col_idx in range(len(df.columns)):
            # Get column label (usually in first row before years)
            label = None
            for row_idx in range(min(year_row_indices)):
                potential_label = df.iloc[row_idx, col_idx]
                if isinstance(potential_label, str):
                    label = potential_label
                    break

            if not label:
                continue

            # Try to match label to standard field
            standard_field = self._fuzzy_match_field(label)
            if not standard_field:
                continue

            # Extract values for each year
            values = []
            for row_idx in year_row_indices:
                value = df.iloc[row_idx, col_idx]
                if isinstance(value, (int, float)):
                    values.append(float(value))
                else:
                    values.append(None)

            metrics[standard_field] = values

        return metrics

    def _fuzzy_match_field(self, label: str, threshold: int = 75) -> Optional[str]:
        """
        Match Excel row/column label to standard field name using fuzzy matching.

        Args:
            label: Label from Excel (e.g., "Net Sales", "SG&A")
            threshold: Minimum fuzzy match score (0-100)

        Returns:
            Standard field name or None if no good match

        Examples:
            "Net Sales" -> "revenue"
            "SG&A" -> "operating_expenses"
            "Depreciation & Amortization" -> "depreciation_amortization"
        """
        label_clean = label.lower().strip()

        # Remove common prefixes/suffixes
        label_clean = re.sub(r'^(total|net|gross)\s+', '', label_clean)
        label_clean = re.sub(r'\s+(expense|income|assets|liabilities)$', '', label_clean)

        best_match = None
        best_score = 0

        for standard_field, synonyms in self.FIELD_SYNONYMS.items():
            # Check exact match first
            if label_clean in synonyms:
                return standard_field

            # Try fuzzy matching
            for synonym in synonyms:
                score = fuzz.ratio(label_clean, synonym)
                if score > best_score and score >= threshold:
                    best_score = score
                    best_match = standard_field

        return best_match

    def _combine_sheet_data(self, all_data: Dict[str, Dict]) -> Dict[str, Any]:
        """
        Combine data from multiple sheets into a single dataset.

        Args:
            all_data: Dict mapping sheet names to extracted data

        Returns:
            Combined data dict
        """
        # Find the most complete set of years
        all_years = []
        for sheet_name, data in all_data.items():
            if data['years']:
                all_years = data['years']
                break

        # Merge metrics from all sheets
        combined_metrics = {}
        for sheet_name, data in all_data.items():
            metrics = data.get('metrics', {})
            for field, values in metrics.items():
                if field not in combined_metrics:
                    combined_metrics[field] = values
                # If field exists in multiple sheets, prefer non-None values
                else:
                    for i, val in enumerate(values):
                        if val is not None and (i >= len(combined_metrics[field]) or combined_metrics[field][i] is None):
                            if i < len(combined_metrics[field]):
                                combined_metrics[field][i] = val

        return {'years': all_years, 'metrics': combined_metrics}

    def _build_financial_data(
        self,
        company_name: str,
        combined_data: Dict,
        source: str
    ) -> FinancialData:
        """
        Build FinancialData object from combined extracted data.

        Args:
            company_name: Name of company
            combined_data: Dict with years and metrics
            source: Source file path

        Returns:
            FinancialData object
        """
        years = combined_data['years']
        metrics = combined_data['metrics']

        if not years:
            raise ValueError("No years found in Excel file")

        num_years = len(years)

        # Helper to get metric values or None list
        def get_metric(field_name):
            return metrics.get(field_name, [None] * num_years)

        # Build income statement
        revenue = get_metric('revenue')
        if not revenue or all(v is None for v in revenue):
            raise ValueError("Revenue not found in Excel file - cannot proceed")

        income_stmt = IncomeStatement(
            revenue=revenue,
            cogs=get_metric('cogs'),
            gross_profit=get_metric('gross_profit'),
            operating_expenses=get_metric('operating_expenses'),
            rd_expense=get_metric('rd_expense'),
            sga_expense=get_metric('sga_expense'),
            ebitda=get_metric('ebitda'),
            depreciation_amortization=get_metric('depreciation_amortization'),
            ebit=get_metric('ebit'),
            interest_expense=get_metric('interest_expense'),
            interest_income=get_metric('interest_income'),
            pretax_income=get_metric('pretax_income'),
            income_tax=get_metric('income_tax'),
            net_income=get_metric('net_income'),
        )

        # Build balance sheet
        balance_sheet = BalanceSheet(
            cash=get_metric('cash'),
            marketable_securities=get_metric('marketable_securities'),
            accounts_receivable=get_metric('accounts_receivable'),
            inventory=get_metric('inventory'),
            current_assets=get_metric('current_assets'),
            ppe_net=get_metric('ppe_net'),
            ppe_gross=get_metric('ppe_gross'),
            goodwill=get_metric('goodwill'),
            intangible_assets=get_metric('intangible_assets'),
            total_assets=get_metric('total_assets'),
            accounts_payable=get_metric('accounts_payable'),
            accrued_expenses=get_metric('accrued_expenses'),
            short_term_debt=get_metric('short_term_debt'),
            current_liabilities=get_metric('current_liabilities'),
            long_term_debt=get_metric('long_term_debt'),
            total_liabilities=get_metric('total_liabilities'),
            shareholders_equity=get_metric('shareholders_equity'),
        )

        # Build cash flow statement
        cash_flow = CashFlowStatement(
            operating_cash_flow=get_metric('operating_cash_flow'),
            depreciation_amortization=get_metric('depreciation_amortization'),
            change_in_nwc=get_metric('change_in_nwc'),
            capex=get_metric('capex'),
            free_cash_flow=get_metric('free_cash_flow'),
        )

        # Market data (usually not in historical Excel, leave empty)
        market_data = MarketData()

        # Metadata
        metadata = ExtractionMetadata(
            source="excel",
            source_path=source,
            extraction_date=datetime.now(),
            notes=f"Extracted from {Path(source).name}"
        )

        # Build financial data
        financial_data = FinancialData(
            company=CompanyInfo(name=company_name),
            years=years,
            income_statement=income_stmt,
            balance_sheet=balance_sheet,
            cash_flow=cash_flow,
            market_data=market_data,
            metadata=metadata,
        )

        # Calculate completeness
        financial_data.metadata.completeness_score = self._calculate_completeness(financial_data)

        # Basic validation
        self._validate_basic_data(financial_data)

        return financial_data
