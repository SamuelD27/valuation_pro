"""
Comprehensive Financial Data Extractor

Extracts data from Financial_Model_Data_Source.xlsx (AcmeTech Holdings)
which contains multi-year financial statements and market data.
"""

import openpyxl
from typing import Dict, List, Optional


class ComprehensiveDataExtractor:
    """
    Extract comprehensive financial data from the multi-tab data source.

    Data source: AcmeTech Holdings Ltd.
    Fiscal year end: June 30
    Years: 2021-2025 (actuals)
    Currency: USD $ millions
    """

    def __init__(self, filepath: str = 'Base_datasource/Financial_Model_Data_Source.xlsx'):
        """
        Initialize extractor.

        Args:
            filepath: Path to Financial_Model_Data_Source.xlsx
        """
        self.filepath = filepath
        self.wb = None

    def load_workbook(self):
        """Load the workbook."""
        if not self.wb:
            self.wb = openpyxl.load_workbook(self.filepath, data_only=True)

    def close_workbook(self):
        """Close the workbook."""
        if self.wb:
            self.wb.close()
            self.wb = None

    def get_income_statement(self) -> Dict:
        """
        Extract Income Statement data.

        Returns:
            Dict with years as keys and financial metrics as values
        """
        self.load_workbook()
        ws = self.wb['Income Statement']

        data = {
            'years': [],
            'revenue': [],
            'cogs': [],
            'gross_profit': [],
            'r_and_d': [],
            'sales_marketing': [],
            'g_and_a': [],
            'stock_based_comp': [],
            'total_opex': [],
            'ebitda_pre_sbc': [],
        }

        # Read data from rows 2-6 (2021-2025)
        for row in range(2, 7):
            data['years'].append(ws.cell(row=row, column=1).value)
            data['revenue'].append(ws.cell(row=row, column=2).value)
            data['cogs'].append(ws.cell(row=row, column=3).value)
            data['gross_profit'].append(ws.cell(row=row, column=4).value)
            data['r_and_d'].append(ws.cell(row=row, column=5).value)
            data['sales_marketing'].append(ws.cell(row=row, column=6).value)
            data['g_and_a'].append(ws.cell(row=row, column=7).value)
            data['stock_based_comp'].append(ws.cell(row=row, column=8).value)
            data['total_opex'].append(ws.cell(row=row, column=9).value)
            data['ebitda_pre_sbc'].append(ws.cell(row=row, column=10).value)

        return data

    def get_balance_sheet(self) -> Dict:
        """
        Extract Balance Sheet data.

        Returns:
            Dict with years as keys and balance sheet items as values
        """
        self.load_workbook()
        ws = self.wb['Balance Sheet']

        data = {
            'years': [],
            'cash': [],
            'accounts_receivable': [],
            'inventory': [],
            'other_current_assets': [],
            'total_current_assets': [],
            'net_ppe': [],
            'intangibles': [],
            'goodwill': [],
            'total_assets': [],
        }

        # Read data from rows 2-6 (2021-2025)
        for row in range(2, 7):
            data['years'].append(ws.cell(row=row, column=1).value)
            data['cash'].append(ws.cell(row=row, column=2).value)
            data['accounts_receivable'].append(ws.cell(row=row, column=3).value)
            data['inventory'].append(ws.cell(row=row, column=4).value)
            data['other_current_assets'].append(ws.cell(row=row, column=5).value)
            data['total_current_assets'].append(ws.cell(row=row, column=6).value)
            data['net_ppe'].append(ws.cell(row=row, column=7).value)
            data['intangibles'].append(ws.cell(row=row, column=8).value)
            data['goodwill'].append(ws.cell(row=row, column=9).value)
            data['total_assets'].append(ws.cell(row=row, column=10).value)

        return data

    def get_cash_flow_statement(self) -> Dict:
        """
        Extract Cash Flow Statement data.

        Returns:
            Dict with years as keys and cash flow items as values
        """
        self.load_workbook()
        ws = self.wb['Cash Flow Statement']

        data = {
            'years': [],
            'net_income': [],
            'd_and_a': [],
            'stock_based_comp': [],
            'change_in_nwc': [],
            'cfo': [],
            'capex': [],
            'acquisitions': [],
            'disposals': [],
            'cfi': [],
        }

        # Read data from rows 2-6 (2021-2025)
        for row in range(2, 7):
            data['years'].append(ws.cell(row=row, column=1).value)
            data['net_income'].append(ws.cell(row=row, column=2).value)
            data['d_and_a'].append(ws.cell(row=row, column=3).value)
            data['stock_based_comp'].append(ws.cell(row=row, column=4).value)
            data['change_in_nwc'].append(ws.cell(row=row, column=5).value)
            data['cfo'].append(ws.cell(row=row, column=6).value)
            data['capex'].append(ws.cell(row=row, column=7).value)
            data['acquisitions'].append(ws.cell(row=row, column=8).value)
            data['disposals'].append(ws.cell(row=row, column=9).value)
            data['cfi'].append(ws.cell(row=row, column=10).value)

        return data

    def get_market_data(self) -> Dict:
        """
        Extract Market Data (for valuation multiples and EV build).

        Returns:
            Dict with market cap, debt, cash, EV data
        """
        self.load_workbook()
        ws = self.wb['Market Data']

        data = {
            'years': [],
            'share_price': [],
            'shares_outstanding': [],
            'market_cap': [],
            'total_debt': [],
            'cash': [],
            'net_debt': [],
            'enterprise_value': [],
        }

        # Read data from rows 2-6 (2021-2025)
        for row in range(2, 7):
            data['years'].append(ws.cell(row=row, column=1).value)
            data['share_price'].append(ws.cell(row=row, column=2).value)
            data['shares_outstanding'].append(ws.cell(row=row, column=3).value)
            data['market_cap'].append(ws.cell(row=row, column=4).value)
            data['total_debt'].append(ws.cell(row=row, column=5).value)
            data['cash'].append(ws.cell(row=row, column=6).value)
            data['net_debt'].append(ws.cell(row=row, column=7).value)
            data['enterprise_value'].append(ws.cell(row=row, column=8).value)

        return data

    def get_ltm_metrics(self) -> Dict:
        """
        Get Last Twelve Months (LTM) metrics for modeling.
        Uses 2025 data (most recent year).

        Returns:
            Dict with key LTM metrics
        """
        is_data = self.get_income_statement()
        bs_data = self.get_balance_sheet()
        cf_data = self.get_cash_flow_statement()
        mkt_data = self.get_market_data()

        # Use 2025 (index -1 = last year)
        ltm = {
            'year': is_data['years'][-1],
            'revenue': is_data['revenue'][-1],
            'ebitda': is_data['ebitda_pre_sbc'][-1],
            'net_income': cf_data['net_income'][-1],
            'd_and_a': cf_data['d_and_a'][-1],
            'capex': cf_data['capex'][-1],
            'total_debt': mkt_data['total_debt'][-1],
            'cash': mkt_data['cash'][-1],
            'net_debt': mkt_data['net_debt'][-1],
            'enterprise_value': mkt_data['enterprise_value'][-1],
            'market_cap': mkt_data['market_cap'][-1],
        }

        return ltm

    def get_historical_data(self, years: int = 5) -> Dict:
        """
        Get historical financial data for specified number of years.

        Args:
            years: Number of historical years (default 5 = all available)

        Returns:
            Dict with complete historical financials
        """
        is_data = self.get_income_statement()
        bs_data = self.get_balance_sheet()
        cf_data = self.get_cash_flow_statement()

        # Return last N years
        historical = {
            'years': is_data['years'][-years:],
            'income_statement': {
                'revenue': is_data['revenue'][-years:],
                'cogs': is_data['cogs'][-years:],
                'gross_profit': is_data['gross_profit'][-years:],
                'ebitda': is_data['ebitda_pre_sbc'][-years:],
                'total_opex': is_data['total_opex'][-years:],
            },
            'balance_sheet': {
                'cash': bs_data['cash'][-years:],
                'total_assets': bs_data['total_assets'][-years:],
                'net_ppe': bs_data['net_ppe'][-years:],
            },
            'cash_flow': {
                'net_income': cf_data['net_income'][-years:],
                'd_and_a': cf_data['d_and_a'][-years:],
                'cfo': cf_data['cfo'][-years:],
                'capex': cf_data['capex'][-years:],
            },
        }

        return historical

    def __enter__(self):
        """Context manager entry."""
        self.load_workbook()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        self.close_workbook()


if __name__ == "__main__":
    # Test extraction
    print("="*80)
    print("TESTING COMPREHENSIVE DATA EXTRACTOR")
    print("="*80)

    with ComprehensiveDataExtractor() as extractor:
        # Test LTM metrics
        ltm = extractor.get_ltm_metrics()
        print("\n📊 LTM METRICS (2025):")
        for key, value in ltm.items():
            print(f"   {key}: {value}")

        # Test historical data
        historical = extractor.get_historical_data(years=3)
        print("\n📈 HISTORICAL DATA (Last 3 Years):")
        print(f"   Years: {historical['years']}")
        print(f"   Revenue: {historical['income_statement']['revenue']}")
        print(f"   EBITDA: {historical['income_statement']['ebitda']}")

    print("\n✅ Extraction successful!")
