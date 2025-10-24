"""
Excel Financial Statement Extractor

Reads financial data from Excel files (Income Statement, Balance Sheet, Cash Flow)
and extracts it into structured format for valuation models.
"""

import openpyxl
from typing import Dict, List, Optional
import pandas as pd
from datetime import datetime


class FinancialStatementExtractor:
    """
    Extract financial data from Excel files.

    Supports standard financial statement formats:
    - Income Statement
    - Balance Sheet
    - Cash Flow Statement
    """

    def __init__(self):
        """Initialize extractor."""
        self.income_statement = None
        self.balance_sheet = None
        self.cash_flow = None

    def extract_income_statement(self, filepath: str) -> Dict:
        """
        Extract income statement data from Excel file.

        Args:
            filepath: Path to income statement Excel file

        Returns:
            Dictionary with income statement data
        """
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Extract years from row 4
        years = []
        year_cols = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=4, column=col).value
            if val and isinstance(val, (int, float)) and val > 2000:
                years.append(int(val))
                year_cols.append(col)

        # Extract revenue data
        revenue = self._extract_line_items(ws, "Revenue", "Total Revenues", year_cols)

        # Extract expense data
        expenses = self._extract_line_items(ws, "Expenses", "Total Expenses", year_cols)

        # Extract COGS
        cogs = self._extract_single_item(ws, "Cost of goods sold", year_cols)

        # Extract depreciation
        depreciation = self._extract_single_item(ws, "Depreciation", year_cols)

        # Extract net income (usually last major line)
        net_income = self._extract_single_item(ws, "Net Income", year_cols)

        self.income_statement = {
            'years': years,
            'revenue': revenue,
            'expenses': expenses,
            'cogs': cogs,
            'depreciation': depreciation,
            'net_income': net_income,
        }

        return self.income_statement

    def extract_balance_sheet(self, filepath: str) -> Dict:
        """
        Extract balance sheet data from Excel file.

        Args:
            filepath: Path to balance sheet Excel file

        Returns:
            Dictionary with balance sheet data
        """
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Extract years from row 4
        years = []
        year_cols = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=4, column=col).value
            if val and isinstance(val, (int, float)) and val > 2000:
                years.append(int(val))
                year_cols.append(col)

        # Extract assets
        cash = self._extract_single_item(ws, "Cash", year_cols)
        accounts_receivable = self._extract_single_item(ws, "Accounts receivable", year_cols)
        inventory = self._extract_single_item(ws, "Inventory", year_cols)
        total_current_assets = self._extract_single_item(ws, "Total current assets", year_cols)

        ppe_gross = self._extract_single_item(ws, "Property, plant, and equipment", year_cols)
        accumulated_dep = self._extract_single_item(ws, "accumulated depreciation", year_cols)
        total_fixed_assets = self._extract_single_item(ws, "Total fixed assets", year_cols)

        total_assets = self._extract_single_item(ws, "Total Assets", year_cols)

        # Extract liabilities
        accounts_payable = self._extract_single_item(ws, "Accounts payable", year_cols)
        total_current_liab = self._extract_single_item(ws, "Total current liabilities", year_cols)
        long_term_debt = self._extract_single_item(ws, "Long-term debt", year_cols)
        total_liabilities = self._extract_single_item(ws, "Total Liabilities", year_cols)

        # Extract equity
        total_equity = self._extract_single_item(ws, "Total Equity", year_cols)

        self.balance_sheet = {
            'years': years,
            'cash': cash,
            'accounts_receivable': accounts_receivable,
            'inventory': inventory,
            'total_current_assets': total_current_assets,
            'ppe_gross': ppe_gross,
            'accumulated_depreciation': accumulated_dep,
            'total_fixed_assets': total_fixed_assets,
            'total_assets': total_assets,
            'accounts_payable': accounts_payable,
            'total_current_liabilities': total_current_liab,
            'long_term_debt': long_term_debt,
            'total_liabilities': total_liabilities,
            'total_equity': total_equity,
        }

        return self.balance_sheet

    def extract_cash_flow_statement(self, filepath: str) -> Dict:
        """
        Extract cash flow statement data from Excel file.

        Args:
            filepath: Path to cash flow statement Excel file

        Returns:
            Dictionary with cash flow data
        """
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Extract year from row 4
        year = None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=4, column=col).value
            if val and isinstance(val, datetime):
                year = val.year
                break

        if not year:
            year = 2019  # Default

        # Extract beginning cash
        beginning_cash = self._extract_single_value(ws, "Cash at Beginning of Year")

        # Extract operating activities
        cash_from_customers = self._extract_single_value(ws, "Customers")
        cash_for_inventory = self._extract_single_value(ws, "Inventory purchases")
        cash_for_expenses = self._extract_single_value(ws, "operating and administrative")
        cash_for_wages = self._extract_single_value(ws, "Wage expenses")
        cash_for_interest = self._extract_single_value(ws, "Interest")
        cash_for_taxes = self._extract_single_value(ws, "Income taxes")

        operating_cf = self._extract_single_value(ws, "Net Cash Flow from Operations")

        # Extract investing activities
        investing_cf = self._extract_single_value(ws, "Net Cash Flow from Investing")

        # Extract financing activities
        financing_cf = self._extract_single_value(ws, "Net Cash Flow from Financing")

        # Net change in cash
        net_change = self._extract_single_value(ws, "Net Change in Cash")

        # Ending cash
        ending_cash = self._extract_single_value(ws, "Cash at End of Year")

        self.cash_flow = {
            'year': year,
            'beginning_cash': beginning_cash or 0,
            'operating_cf': operating_cf or 0,
            'investing_cf': investing_cf or 0,
            'financing_cf': financing_cf or 0,
            'net_change': net_change or 0,
            'ending_cash': ending_cash or 0,
        }

        return self.cash_flow

    def _extract_line_items(
        self,
        ws,
        start_keyword: str,
        total_keyword: str,
        year_cols: List[int]
    ) -> Dict[str, List[float]]:
        """
        Extract multiple line items between start and total keywords.

        Args:
            ws: Worksheet
            start_keyword: Keyword marking start of section
            total_keyword: Keyword marking total/end of section
            year_cols: List of column indices for years

        Returns:
            Dictionary of line items with values by year
        """
        items = {}
        in_section = False

        for row in range(1, ws.max_row + 1):
            label = ws.cell(row=row, column=2).value
            if not label:
                label = ws.cell(row=row, column=1).value

            if not label:
                continue

            label = str(label).strip()

            if start_keyword.lower() in label.lower():
                in_section = True
                continue

            if total_keyword.lower() in label.lower():
                in_section = False
                # Also capture the total
                values = []
                for col in year_cols:
                    val = ws.cell(row=row, column=col).value
                    if val and not isinstance(val, str):
                        values.append(float(val))
                    else:
                        values.append(0.0)
                items[label] = values
                break

            if in_section and label:
                values = []
                for col in year_cols:
                    val = ws.cell(row=row, column=col).value
                    if val and not isinstance(val, str):
                        values.append(float(val))
                    else:
                        values.append(0.0)
                if any(v != 0 for v in values):
                    items[label] = values

        return items

    def _extract_single_item(self, ws, keyword: str, year_cols: List[int]) -> List[float]:
        """
        Extract values for a single line item across years.

        Args:
            ws: Worksheet
            keyword: Keyword to search for
            year_cols: List of column indices for years

        Returns:
            List of values across years
        """
        for row in range(1, ws.max_row + 1):
            for col in range(1, 4):  # Check first 3 columns for label
                label = ws.cell(row=row, column=col).value
                if label and keyword.lower() in str(label).lower():
                    values = []
                    for year_col in year_cols:
                        val = ws.cell(row=row, column=year_col).value
                        if val and not isinstance(val, str):
                            values.append(float(val))
                        else:
                            values.append(0.0)
                    return values

        return [0.0] * len(year_cols)

    def _extract_single_value(self, ws, keyword: str) -> Optional[float]:
        """
        Extract a single value from worksheet.

        Args:
            ws: Worksheet
            keyword: Keyword to search for

        Returns:
            Value or None
        """
        for row in range(1, ws.max_row + 1):
            for col in range(1, 5):
                label = ws.cell(row=row, column=col).value
                if label and keyword.lower() in str(label).lower():
                    # Look for value in next columns
                    for val_col in range(col + 1, min(col + 4, ws.max_column + 1)):
                        val = ws.cell(row=row, column=val_col).value
                        if val and not isinstance(val, str):
                            try:
                                return float(val)
                            except:
                                pass
        return None

    def get_summary(self) -> Dict:
        """
        Get summary of all extracted financial data.

        Returns:
            Dictionary with all financial data
        """
        return {
            'income_statement': self.income_statement,
            'balance_sheet': self.balance_sheet,
            'cash_flow': self.cash_flow,
        }
