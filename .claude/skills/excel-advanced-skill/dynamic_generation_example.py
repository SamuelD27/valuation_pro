"""
Dynamic Excel Model Generation - Complete Working Example

Demonstrates: Input analysis → Capability mapping → Component selection → Model generation
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from typing import Dict


class InputAnalyzer:
    """Analyzes inputs and determines what can be generated"""
    
    def analyze_inputs(self, data: Dict) -> Dict:
        """Returns capability map showing feasibility"""
        capabilities = {'analyses': {}, 'components': [], 'quality_score': 0.0}
        
        # Check DCF
        dcf_fields = ['revenue', 'ebitda', 'tax_rate', 'capex']
        if all(f in data for f in dcf_fields):
            capabilities['analyses']['dcf'] = {'feasible': True, 'level': 'basic'}
            capabilities['components'].append('fcf_projection')
        
        # Check LBO
        lbo_fields = ['ebitda', 'enterprise_value', 'debt']
        if all(f in data for f in lbo_fields):
            capabilities['analyses']['lbo'] = {'feasible': True}
            capabilities['components'].append('lbo_returns')
        
        # Calculate quality score
        feasible = sum(1 for a in capabilities['analyses'].values() if a.get('feasible'))
        capabilities['quality_score'] = feasible / 2  # DCF + LBO
        
        return capabilities


class FormulaBuilder:
    """Generates Excel formulas with proper references"""
    
    def fcf_formula(self, ebitda, da, tax, capex, nwc):
        """FCF = NOPAT - CapEx - ΔNW"""
        nopat = f"(({ebitda}-{da})*(1-{tax}))"
        return f"={nopat}-{capex}-{nwc}"
    
    def wacc_formula(self, equity, debt, re, rd, tax):
        """WACC = (E/V)*Re + (D/V)*Rd*(1-T)"""
        v = f"({equity}+{debt})"
        return f"=({equity}/{v})*{re}+({debt}/{v})*{rd}*(1-{tax})"


class DynamicModelGenerator:
    """Orchestrates dynamic model generation"""
    
    def __init__(self):
        self.analyzer = InputAnalyzer()
        self.formula_builder = FormulaBuilder()
    
    def generate(self, data: Dict, company: str) -> Workbook:
        """Generate custom model based on available inputs"""
        
        # Step 1: Analyze
        caps = self.analyzer.analyze_inputs(data)
        print(f"\nGenerating {company} model:")
        print(f"  Quality: {caps['quality_score']:.0%}")
        print(f"  Components: {caps['components']}")
        
        # Step 2: Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{company} Valuation"
        
        # Step 3: Generate assumptions
        ws['A1'] = "Revenue Growth"
        ws['B1'] = data.get('revenue_growth', 0.10)
        ws['B1'].number_format = '0.0%'
        
        ws['A2'] = "Tax Rate"
        ws['B2'] = data.get('tax_rate', 0.25)
        ws['B2'].number_format = '0.0%'
        
        # Step 4: Generate DCF if feasible
        row = 4
        if caps['analyses'].get('dcf', {}).get('feasible'):
            ws[f'A{row}'] = "DCF VALUATION"
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = "Revenue"
            ws[f'B{row}'] = data['revenue']
            ws[f'C{row}'] = f"=B{row}*(1+$B$1)"  # Growth formula with absolute reference
            
        return wb


# Example usage
if __name__ == "__main__":
    
    # Example 1: Full data
    data_full = {
        'company': 'TechCorp',
        'revenue': 1500,
        'ebitda': 450,
        'capex': 120,
        'tax_rate': 0.25,
        'revenue_growth': 0.15,
        'enterprise_value': 8000,
        'debt': 500,
    }
    
    generator = DynamicModelGenerator()
    wb1 = generator.generate(data_full, 'TechCorp')
    wb1.save('/tmp/techcorp_model.xlsx')
    
    # Example 2: Limited data
    data_limited = {
        'company': 'StartupCo',
        'revenue': 500,
        'ebitda': 100,
        'capex': 40,
        'tax_rate': 0.25,
    }
    
    wb2 = generator.generate(data_limited, 'StartupCo')
    wb2.save('/tmp/startupco_model.xlsx')
    
    print("\n✅ Models generated in /tmp/")
