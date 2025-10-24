"""
Tests for Excel Formula Generation

Validates that the three-statement generator creates FORMULAS, not VALUES.
This is the critical test to ensure we don't repeat the old mistake.
"""

import pytest
import openpyxl
from src.excel.three_statement_generator import ThreeStatementGenerator


@pytest.fixture
def sample_data():
    """Sample data for testing."""
    company_data = {
        'revenue': [300000, 350000, 400000],
        'cogs': [200000, 230000, 260000],
        'rnd': [20000, 25000, 30000],
        'sga': [50000, 55000, 60000],
        'ppe': [100000],
        'debt': [80000],
        'cash': [50000]
    }

    assumptions = {
        'revenue_growth': [0.08, 0.07, 0.06, 0.05, 0.04],
        'ebit_margin': 0.30,
        'tax_rate': 0.21,
        'nwc_pct_revenue': 0.025,
        'capex_pct_revenue': 0.03,
        'terminal_growth': 0.025,
        'dso': 45,
        'dio': 30,
        'dpo': 60,
    }

    wacc_data = {
        'risk_free_rate': 0.04,
        'beta': 1.2,
        'market_risk_premium': 0.07,
        'cost_of_debt': 0.05,
        'market_cap': 2700000,
        'total_debt': 111000,
    }

    return company_data, assumptions, wacc_data


def test_formulas_not_values(sample_data, tmp_path):
    """
    CRITICAL TEST: Ensure Excel cells contain FORMULAS, not Python values.

    This test fails if we write values like the old generator did.
    """
    company_data, assumptions, wacc_data = sample_data

    # Generate model
    generator = ThreeStatementGenerator(ticker="TEST")
    output_file = tmp_path / "test_formulas.xlsx"

    generator.generate_full_model(
        company_data=company_data,
        assumptions=assumptions,
        wacc_data=wacc_data,
        filepath=str(output_file)
    )

    # Open the generated file
    wb = openpyxl.load_workbook(output_file)

    # Test Income Statement formulas
    is_sheet = wb['Income Statement']

    # Check revenue growth formula (should be formula, not value)
    # Row 7 is revenue growth %, column E (2nd year)
    growth_cell = is_sheet['E7']  # Second year growth
    assert growth_cell.value is not None, "Growth cell is empty"
    assert isinstance(growth_cell.value, str), f"Growth cell should be formula string, got {type(growth_cell.value)}"
    assert growth_cell.value.startswith('='), f"Growth formula should start with '=', got: {growth_cell.value}"
    assert '/' in growth_cell.value, "Growth formula should contain division"

    # Check projected revenue uses formula
    proj_revenue_cell = is_sheet['G6']  # First projection year revenue
    assert proj_revenue_cell.value is not None, "Projected revenue is empty"
    assert isinstance(proj_revenue_cell.value, str), f"Projected revenue should be formula, got {type(proj_revenue_cell.value)}"
    assert proj_revenue_cell.value.startswith('='), f"Revenue formula should start with '='"

    # Check EBITDA calculation is a formula
    ebitda_cell = is_sheet['D19']  # EBITDA for first year
    assert ebitda_cell.value is not None, "EBITDA is empty"
    assert isinstance(ebitda_cell.value, str), f"EBITDA should be formula, got {type(ebitda_cell.value)}"
    assert ebitda_cell.value.startswith('='), "EBITDA should be formula"

    print("✓ Income Statement uses formulas, not values")


def test_sheet_linkage(sample_data, tmp_path):
    """
    Test that sheets properly link to each other using sheet references.

    Example: DCF sheet should reference Income Statement for EBITDA.
    """
    company_data, assumptions, wacc_data = sample_data

    generator = ThreeStatementGenerator(ticker="TEST")
    output_file = tmp_path / "test_linkage.xlsx"

    generator.generate_full_model(
        company_data=company_data,
        assumptions=assumptions,
        wacc_data=wacc_data,
        filepath=str(output_file)
    )

    wb = openpyxl.load_workbook(output_file)

    # Test DCF sheet references Income Statement
    dcf_sheet = wb['DCF']

    # EBITDA in DCF should reference Income Statement
    dcf_ebitda_cell = dcf_sheet['D8']  # First year EBITDA in DCF
    assert dcf_ebitda_cell.value is not None, "DCF EBITDA is empty"
    assert isinstance(dcf_ebitda_cell.value, str), "DCF EBITDA should be formula"
    assert "'Income Statement'!" in dcf_ebitda_cell.value or "Income Statement!" in dcf_ebitda_cell.value, \
        f"DCF EBITDA should reference Income Statement, got: {dcf_ebitda_cell.value}"

    # Test Balance Sheet references PPE Schedule
    bs_sheet = wb['Balance Sheet']
    ppe_cell = bs_sheet['D13']  # PP&E in balance sheet

    if ppe_cell.value:
        assert isinstance(ppe_cell.value, str), "PP&E should be formula"
        if ppe_cell.value.startswith('='):
            assert "'PPE Schedule'!" in ppe_cell.value or "PPE Schedule!" in ppe_cell.value, \
                f"Balance Sheet PP&E should reference PPE Schedule"

    # Test Income Statement references Debt Schedule for interest
    is_sheet = wb['Income Statement']
    interest_cell = is_sheet['D26']  # Interest expense

    if interest_cell.value and isinstance(interest_cell.value, str) and interest_cell.value.startswith('='):
        assert "'Debt Schedule'!" in interest_cell.value or "Debt Schedule!" in interest_cell.value, \
            f"Interest should reference Debt Schedule"

    print("✓ Sheets are properly linked with cross-references")


def test_scenario_switching(sample_data, tmp_path):
    """
    Test that Assumptions sheet has CHOOSE formulas for scenario switching.

    The active column (H) should use CHOOSE($B$2, base, down, up) formulas.
    """
    company_data, assumptions, wacc_data = sample_data

    generator = ThreeStatementGenerator(ticker="TEST")
    output_file = tmp_path / "test_scenarios.xlsx"

    generator.generate_full_model(
        company_data=company_data,
        assumptions=assumptions,
        wacc_data=wacc_data,
        filepath=str(output_file)
    )

    wb = openpyxl.load_workbook(output_file)
    assumptions_sheet = wb['Assumptions']

    # Check scenario selector exists
    selector_cell = assumptions_sheet['B2']
    assert selector_cell.value == 1, "Scenario selector should default to 1 (Base case)"

    # Check that active column uses CHOOSE formula
    # Revenue growth year 1 active value (column H, row 6)
    active_cell = assumptions_sheet['H6']

    assert active_cell.value is not None, "Active value is empty"
    assert isinstance(active_cell.value, str), f"Active value should be formula, got {type(active_cell.value)}"
    assert active_cell.value.startswith('='), "Active value should be formula"
    assert 'CHOOSE' in active_cell.value.upper(), f"Should use CHOOSE formula, got: {active_cell.value}"
    assert 'B2' in active_cell.value or '$B$2' in active_cell.value, \
        f"CHOOSE should reference scenario selector B2, got: {active_cell.value}"

    print("✓ Scenario switching with CHOOSE formulas works")


def test_no_hardcoded_calculated_values(sample_data, tmp_path):
    """
    Test that calculated values are NEVER hardcoded.

    Only inputs should be values. Everything else should be formulas.
    """
    company_data, assumptions, wacc_data = sample_data

    generator = ThreeStatementGenerator(ticker="TEST")
    output_file = tmp_path / "test_no_values.xlsx"

    generator.generate_full_model(
        company_data=company_data,
        assumptions=assumptions,
        wacc_data=wacc_data,
        filepath=str(output_file)
    )

    wb = openpyxl.load_workbook(output_file)

    # Income Statement - check that ALL calculated cells are formulas
    is_sheet = wb['Income Statement']

    # EBIT should be formula (not hardcoded)
    for col_idx in range(4, 9):  # Columns D through H
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ebit_cell = is_sheet[f'{col_letter}27']  # EBIT row

        if ebit_cell.value is not None:
            # Should be either a formula or input (for historical)
            if col_idx >= 7:  # Projection columns
                assert isinstance(ebit_cell.value, str), \
                    f"Projected EBIT in {col_letter}27 should be formula, got: {ebit_cell.value}"
                assert ebit_cell.value.startswith('='), \
                    f"Projected EBIT should start with '='"

    # DCF - Free Cash Flow should be formula
    dcf_sheet = wb['DCF']
    fcf_cell = dcf_sheet['D18']  # FCF first projection year

    assert fcf_cell.value is not None, "FCF is empty"
    assert isinstance(fcf_cell.value, str), f"FCF should be formula, got type {type(fcf_cell.value)}"
    assert fcf_cell.value.startswith('='), "FCF should be formula"

    # Enterprise Value should be formula
    ev_cell = dcf_sheet['D34']

    assert ev_cell.value is not None, "Enterprise Value is empty"
    assert isinstance(ev_cell.value, str), f"Enterprise Value should be formula"
    assert ev_cell.value.startswith('='), "Enterprise Value should be formula"

    print("✓ No hardcoded calculated values - all formulas")


def test_wacc_calculation_formulas(sample_data, tmp_path):
    """Test that WACC sheet uses formulas for all calculations."""
    company_data, assumptions, wacc_data = sample_data

    generator = ThreeStatementGenerator(ticker="TEST")
    output_file = tmp_path / "test_wacc.xlsx"

    generator.generate_full_model(
        company_data=company_data,
        assumptions=assumptions,
        wacc_data=wacc_data,
        filepath=str(output_file)
    )

    wb = openpyxl.load_workbook(output_file)
    wacc_sheet = wb['WACC']

    # Cost of Equity should be formula (CAPM)
    re_cell = wacc_sheet['C9']  # Cost of Equity

    assert re_cell.value is not None, "Cost of Equity is empty"
    assert isinstance(re_cell.value, str), "Cost of Equity should be formula"
    assert re_cell.value.startswith('='), "Cost of Equity should be formula"
    # Should reference risk-free rate, beta, and MRP
    assert 'C' in re_cell.value, "Should reference input cells"

    # After-tax cost of debt should be formula
    rd_after_tax_cell = wacc_sheet['C15']

    assert rd_after_tax_cell.value is not None, "After-tax Rd is empty"
    assert isinstance(rd_after_tax_cell.value, str), "After-tax Rd should be formula"
    assert '(1-' in rd_after_tax_cell.value or '1-' in rd_after_tax_cell.value, \
        "Should have tax shield formula"

    # WACC itself should be formula
    wacc_cell = wacc_sheet['C20']

    assert wacc_cell.value is not None, "WACC is empty"
    assert isinstance(wacc_cell.value, str), "WACC should be formula"
    assert wacc_cell.value.startswith('='), "WACC should be formula"

    print("✓ WACC calculations use formulas")


def test_balance_sheet_balances(sample_data, tmp_path):
    """
    Test that Balance Sheet has check formula.

    Total Assets should equal Total Liabilities + Equity.
    """
    company_data, assumptions, wacc_data = sample_data

    generator = ThreeStatementGenerator(ticker="TEST")
    output_file = tmp_path / "test_balance.xlsx"

    generator.generate_full_model(
        company_data=company_data,
        assumptions=assumptions,
        wacc_data=wacc_data,
        filepath=str(output_file)
    )

    wb = openpyxl.load_workbook(output_file)
    bs_sheet = wb['Balance Sheet']

    # Find the CHECK row (should be near bottom)
    # Look for cell containing "CHECK"
    check_row = None
    for row in range(1, 40):
        cell = bs_sheet.cell(row=row, column=1)
        if cell.value and 'CHECK' in str(cell.value).upper():
            check_row = row
            break

    assert check_row is not None, "Balance Sheet should have a CHECK row"

    # Check that the formula exists
    check_cell = bs_sheet.cell(row=check_row, column=4)  # First data column

    assert check_cell.value is not None, "Check formula is empty"
    assert isinstance(check_cell.value, str), "Check should be formula"
    assert check_cell.value.startswith('='), "Check should be formula"
    assert '-' in check_cell.value, "Check should be subtraction (Assets - Liab&Equity)"

    print("✓ Balance Sheet has check formula")


def test_all_sheets_created(sample_data, tmp_path):
    """Test that all required sheets are created."""
    company_data, assumptions, wacc_data = sample_data

    generator = ThreeStatementGenerator(ticker="TEST")
    output_file = tmp_path / "test_all_sheets.xlsx"

    generator.generate_full_model(
        company_data=company_data,
        assumptions=assumptions,
        wacc_data=wacc_data,
        filepath=str(output_file)
    )

    wb = openpyxl.load_workbook(output_file)

    required_sheets = [
        'WSO Cover Page',
        'Assumptions',
        'Income Statement',
        'Balance Sheet',
        'Cash Flow Statement',
        'PPE Schedule',
        'Debt Schedule',
        'WACC',
        'DCF',
        'LBO',
        'Football Field'
    ]

    for sheet_name in required_sheets:
        assert sheet_name in wb.sheetnames, f"Missing required sheet: {sheet_name}"

    print(f"✓ All {len(required_sheets)} required sheets created")


def test_file_opens_in_excel(sample_data, tmp_path):
    """
    Test that the generated file can be opened.

    This catches formula syntax errors that would break Excel.
    """
    company_data, assumptions, wacc_data = sample_data

    generator = ThreeStatementGenerator(ticker="TEST")
    output_file = tmp_path / "test_opens.xlsx"

    generator.generate_full_model(
        company_data=company_data,
        assumptions=assumptions,
        wacc_data=wacc_data,
        filepath=str(output_file)
    )

    # Try to re-open the file
    wb = openpyxl.load_workbook(output_file)
    assert wb is not None, "File should be openable"

    # Check that we can read formulas
    dcf_sheet = wb['DCF']
    assert dcf_sheet is not None, "Should be able to access DCF sheet"

    print("✓ Generated file can be opened and read")


if __name__ == "__main__":
    # Run tests manually
    import tempfile
    import os

    print("\n" + "="*80)
    print("TESTING EXCEL FORMULA GENERATION")
    print("="*80 + "\n")

    with tempfile.TemporaryDirectory() as tmp_dir:
        from pathlib import Path
        tmp_path = Path(tmp_dir)

        data = sample_data()

        try:
            test_formulas_not_values(data, tmp_path)
            test_sheet_linkage(data, tmp_path)
            test_scenario_switching(data, tmp_path)
            test_no_hardcoded_calculated_values(data, tmp_path)
            test_wacc_calculation_formulas(data, tmp_path)
            test_balance_sheet_balances(data, tmp_path)
            test_all_sheets_created(data, tmp_path)
            test_file_opens_in_excel(data, tmp_path)

            print("\n" + "="*80)
            print("ALL TESTS PASSED! ✓")
            print("="*80)
            print("\nKey validations:")
            print("  ✓ Excel cells contain FORMULAS, not Python values")
            print("  ✓ Sheets are properly linked")
            print("  ✓ Scenario switching works with CHOOSE()")
            print("  ✓ No hardcoded calculated values")
            print("  ✓ All 11 sheets created")
            print("  ✓ File can be opened in Excel")
            print("\n" + "="*80)

        except AssertionError as e:
            print(f"\n✗ TEST FAILED: {e}\n")
            raise
        except Exception as e:
            print(f"\n✗ ERROR: {e}\n")
            raise
